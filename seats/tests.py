from django.test import TestCase, TransactionTestCase, override_settings, Client
from django.urls import reverse
from django.core.files.uploadedfile import SimpleUploadedFile
from django.utils import timezone
from datetime import timedelta
import json
import importlib.util
import unittest
import urllib.error
import uuid
import zipfile
import ssl
import tempfile
from io import BytesIO
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import patch, Mock

import httpx
import openai
import openpyxl
import pandas as pd
import desktop_runtime
from seats import cloud as cloud_module

from desktop_shell import (
    DesktopBridge,
    build_multipart_form_data,
    build_file_dialog_types,
    ensure_allowed_extension,
    is_allowed_extension,
    normalize_accept_extensions,
    parse_content_disposition_filename,
    resolve_local_export_url,
)
from seats.cloud import CloudAPIError
from seats.crypto import decrypt_payload, encrypt_payload, generate_rsa_keypair
from seats.data_sharing import DATA_SHARING_ENABLED_KEY, DATA_SHARING_PROMPT_SEEN_VERSION_KEY
from .models import (
    Classroom,
    FrontendKVStore,
    SeatConstraint,
    SeatCellType,
    SeatGroup,
    FutureModeConfig,
    AIConversation,
    AIConversationMessage,
    Seat,
    Student,
    StudentTag,
    StudentTagMembership,
    StudentTagRule,
    LayoutSnapshot,
    ClassroomHistoryEntry,
    SyncMeta,
    CloudSession,
    OnboardingState,
    ONBOARDING_SEEN_STORE_KEY,
    ONBOARDING_SEEN_STORE_VALUE,
)
from .plugin_system import plugin_registry
from .views import (
    APP_MANIFEST_REDIRECT_URL,
    UPDATE_DETAILS_REDIRECT_URL,
    _arrange_standard,
    _arrange_grouped,
    _apply_internal_policy,
    _process_import,
    IMPORT_MODE_MATCH,
    IMPORT_MODE_REPLACE,
    _bsce_json_post,
    _create_future_mode_response,
    _run_future_mode_chat,
    ONBOARDING_SAMPLE_NAME,
)


class LayoutMirrorTests(TestCase):
    def setUp(self):
        self.classroom = Classroom.objects.create(name="镜像测试班", rows=2, cols=4)
        self.group = SeatGroup.objects.create(classroom=self.classroom, name="第一组", order=1)
        self.student_left = Student.objects.create(classroom=self.classroom, name="左侧学生", student_id="S001", score=88)
        self.student_right = Student.objects.create(classroom=self.classroom, name="右侧学生", student_id="S002", score=92)

        self.seat_11 = Seat.objects.get(classroom=self.classroom, row=1, col=1)
        self.seat_12 = Seat.objects.get(classroom=self.classroom, row=1, col=2)
        self.seat_13 = Seat.objects.get(classroom=self.classroom, row=1, col=3)
        self.seat_14 = Seat.objects.get(classroom=self.classroom, row=1, col=4)
        self.seat_11.student = self.student_left
        self.seat_11.group = self.group
        self.seat_11.cell_type = SeatCellType.SEAT
        self.seat_11.save()
        self.seat_12.cell_type = SeatCellType.AISLE
        self.seat_12.save(update_fields=["cell_type"])
        self.seat_13.student = self.student_right
        self.seat_13.group = self.group
        self.seat_13.cell_type = SeatCellType.SEAT
        self.seat_13.save()
        self.seat_14.cell_type = SeatCellType.PODIUM
        self.seat_14.save(update_fields=["cell_type"])

        SeatConstraint.objects.create(
            classroom=self.classroom,
            constraint_type=SeatConstraint.ConstraintType.MUST_SEAT,
            student=self.student_left,
            row=1,
            col=1,
            distance=1,
            enabled=True,
        )

    def test_mirror_layout_flips_students_cells_and_constraints(self):
        url = reverse("mirror_layout", args=[self.classroom.pk])
        response = self.client.post(
            url,
            data=json.dumps({"axis": "lr"}),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload.get("status"), "success")
        self.assertEqual(payload.get("mirror_axis"), "lr")

        seat_map = {
            (seat.row, seat.col): seat
            for seat in Seat.objects.filter(classroom=self.classroom).select_related("student", "group")
        }
        self.assertEqual(seat_map[(1, 4)].student_id, self.student_left.pk)
        self.assertEqual(seat_map[(1, 4)].group_id, self.group.pk)
        self.assertEqual(seat_map[(1, 3)].cell_type, SeatCellType.AISLE)
        self.assertEqual(seat_map[(1, 2)].student_id, self.student_right.pk)
        self.assertEqual(seat_map[(1, 1)].cell_type, SeatCellType.PODIUM)

        constraint = SeatConstraint.objects.get(classroom=self.classroom, student=self.student_left)
        self.assertEqual(constraint.row, 1)
        self.assertEqual(constraint.col, 4)


class ClassroomUnseatedStudentTests(TestCase):
    def setUp(self):
        self.classroom = Classroom.objects.create(name="未入座测试班", rows=2, cols=2)
        self.seated_student = Student.objects.create(classroom=self.classroom, name="已入座学生", student_id="U001")
        self.unseated_student = Student.objects.create(classroom=self.classroom, name="未入座学生", student_id="U002")
        self.seat = Seat.objects.get(classroom=self.classroom, row=1, col=1)
        self.seat.student = self.seated_student
        self.seat.save(update_fields=["student"])

    def test_classroom_detail_only_lists_truly_unseated_students(self):
        response = self.client.get(reverse("classroom_detail", args=[self.classroom.pk]))

        self.assertEqual(response.status_code, 200)
        unseated_students = list(response.context["unseated_students"])
        self.assertEqual([student.pk for student in unseated_students], [self.unseated_student.pk])

    def test_delete_student_rejects_seated_student(self):
        response = self.client.post(
            reverse("delete_student", args=[self.classroom.pk, self.seated_student.pk]),
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 400)
        payload = response.json()
        self.assertEqual(payload.get("status"), "error")
        self.assertIn("已入座", payload.get("message", ""))
        self.assertTrue(Student.objects.filter(pk=self.seated_student.pk).exists())
        self.seat.refresh_from_db()
        self.assertEqual(self.seat.student_id, self.seated_student.pk)


class ClassroomDeleteTests(TransactionTestCase):
    def create_cloud_session(self):
        return CloudSession.objects.create(
            uid="u-delete",
            nickname="老三",
            session_token="delete-token",
            token_expires_at=timezone.now() + timedelta(days=1),
        )

    def test_delete_classroom_suspends_sync_bumps_during_cascade(self):
        classroom = Classroom.objects.create(name="删除同步班", rows=2, cols=2)
        student = Student.objects.create(classroom=classroom, name="同步学生", student_id="D001")
        group = SeatGroup.objects.create(classroom=classroom, name="第一组", leader=student)
        seat = Seat.objects.get(classroom=classroom, row=1, col=1)
        seat.student = student
        seat.group = group
        seat.save(update_fields=["student", "group"])
        classroom.left_guardian = student
        classroom.save(update_fields=["left_guardian"])
        classroom_id = classroom.pk
        sync_meta_id = classroom.sync_meta.pk

        response = self.client.post(reverse("delete_classroom", args=[classroom_id]))

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, reverse("index"))
        self.assertFalse(Classroom.objects.filter(pk=classroom_id).exists())
        self.assertFalse(SyncMeta.objects.filter(pk=sync_meta_id).exists())
        self.assertFalse(Seat.objects.filter(classroom_id=classroom_id).exists())
        self.assertFalse(Student.objects.filter(classroom_id=classroom_id).exists())
        self.assertFalse(SeatGroup.objects.filter(classroom_id=classroom_id).exists())

    def test_delete_backed_up_classroom_deletes_cloud_record_first(self):
        classroom = Classroom.objects.create(name="云端删除班", rows=2, cols=2)
        meta = SyncMeta.objects.get(classroom=classroom)
        meta.cloud_version = 2
        meta.local_version = 2
        meta.last_sync_at = timezone.now()
        meta.save(update_fields=["cloud_version", "local_version", "last_sync_at", "updated_at"])
        classroom_id = classroom.pk
        self.create_cloud_session()

        with patch("seats.views.cloud_api_request") as cloud_request:
            cloud_request.return_value = {"ok": True, "status": "success", "uuid": str(meta.uuid), "version": 3}
            response = self.client.post(reverse("delete_classroom", args=[classroom_id]))

        self.assertEqual(response.status_code, 302)
        self.assertFalse(Classroom.objects.filter(pk=classroom_id).exists())
        cloud_request.assert_called_once()
        self.assertEqual(cloud_request.call_args.args[1:3], ("DELETE", f"/api/sync/{meta.uuid}"))
        delete_body = cloud_request.call_args.args[3]
        self.assertEqual(delete_body["base_version"], 2)
        self.assertEqual(delete_body["device_id"], "local-delete")

    def test_delete_backed_up_classroom_requires_cloud_session(self):
        classroom = Classroom.objects.create(name="未登录删除班", rows=1, cols=1)
        meta = SyncMeta.objects.get(classroom=classroom)
        meta.cloud_version = 1
        meta.local_version = 1
        meta.last_sync_at = timezone.now()
        meta.save(update_fields=["cloud_version", "local_version", "last_sync_at", "updated_at"])

        response = self.client.post(reverse("delete_classroom", args=[classroom.pk]))

        self.assertEqual(response.status_code, 401)
        self.assertTrue(Classroom.objects.filter(pk=classroom.pk).exists())
        self.assertIn("请先登录云服务", response.json()["message"])

    def test_delete_backed_up_classroom_ignores_missing_cloud_record(self):
        classroom = Classroom.objects.create(name="云端已删班", rows=1, cols=1)
        meta = SyncMeta.objects.get(classroom=classroom)
        meta.cloud_version = 1
        meta.local_version = 1
        meta.last_sync_at = timezone.now()
        meta.save(update_fields=["cloud_version", "local_version", "last_sync_at", "updated_at"])
        classroom_id = classroom.pk
        self.create_cloud_session()

        with patch("seats.views.cloud_api_request") as cloud_request:
            cloud_request.side_effect = CloudAPIError(
                "班级不存在",
                status_code=404,
                payload={"status": "error", "message": "班级不存在"},
            )
            response = self.client.post(reverse("delete_classroom", args=[classroom_id]))

        self.assertEqual(response.status_code, 302)
        self.assertFalse(Classroom.objects.filter(pk=classroom_id).exists())


class BsceCloudImportTests(TestCase):
    def setUp(self):
        self.classroom = Classroom.objects.create(name="BSCE云导入测试班", rows=2, cols=2)

    def test_bsce_json_post_randomizes_csrf_cookie_per_request(self):
        class FakeResponse:
            def __enter__(self):
                return self

            def __exit__(self, exc_type, exc, tb):
                return False

            def read(self):
                return b'{"success": true, "data": {}}'

        captured_requests = []

        def fake_urlopen(req, timeout=None, context=None):
            captured_requests.append(req)
            return FakeResponse()

        with patch("seats.views.secrets.token_hex", side_effect=["a" * 40, "b" * 40]):
            with patch("seats.views.urllib.request.urlopen", side_effect=fake_urlopen):
                _bsce_json_post("https://sce.jbyc.cc/api/auth.php", {"action": "login"})
                _bsce_json_post("https://sce.jbyc.cc/api/auth.php", {"action": "login"})

        self.assertEqual(len(captured_requests), 2)
        first_body = json.loads(captured_requests[0].data.decode("utf-8"))
        second_body = json.loads(captured_requests[1].data.decode("utf-8"))

        self.assertEqual(first_body.get("_csrf"), "a" * 40)
        self.assertEqual(second_body.get("_csrf"), "b" * 40)
        self.assertNotEqual(first_body.get("_csrf"), second_body.get("_csrf"))
        self.assertEqual(captured_requests[0].get_header("X-csrf-token"), "a" * 40)
        self.assertEqual(captured_requests[1].get_header("X-csrf-token"), "b" * 40)


class CloudCryptoClientTests(TestCase):
    def test_save_cloud_session_persists_local_and_server_keys(self):
        local_keys = cloud_module.get_or_create_local_cloud_keypair()
        server_keys = generate_rsa_keypair()

        session = cloud_module.save_cloud_session_from_payload({
            'uid': 'u-crypto',
            'nickname': '老三',
            'session_token': 'token-crypto',
            'token_expires_at': (timezone.now() + timedelta(days=1)).isoformat(),
            'client_key_id': local_keys['key_id'],
            'server_key': {
                'key_id': server_keys['key_id'],
                'public_key': server_keys['public_key_pem'],
            },
            'subscription': {'tier': 'free', 'display_name': '免费版', 'limits': {}},
        })

        self.assertEqual(session.client_key_id, local_keys['key_id'])
        self.assertTrue(session.client_private_key_pem)
        self.assertEqual(session.server_key_id, server_keys['key_id'])
        self.assertEqual(session.server_public_key_pem, server_keys['public_key_pem'])

    def test_cloud_api_request_decrypts_encrypted_response(self):
        client_keys = generate_rsa_keypair()
        session = CloudSession.objects.create(
            uid='u-e2ee',
            nickname='老三',
            session_token='token-e2ee',
            client_key_id=client_keys['key_id'],
            client_public_key_pem=client_keys['public_key_pem'],
            client_private_key_pem=client_keys['private_key_pem'],
            server_key_id='server-key',
            server_public_key_pem=generate_rsa_keypair()['public_key_pem'],
            token_expires_at=timezone.now() + timedelta(days=1),
        )

        encrypted_response = {
            'ok': True,
            'status': 'success',
            'encrypted': encrypt_payload({'status': 'success', 'value': 42}, client_keys['public_key_pem'], sender_key_id='server-key'),
        }
        with patch('seats.cloud._request_json', return_value=encrypted_response):
            payload = cloud_module.cloud_api_request(session, 'GET', '/api/me')

        self.assertEqual(payload['status'], 'success')
        self.assertEqual(payload['value'], 42)

    def test_cloud_api_request_refreshes_subscription_before_server_call(self):
        session = CloudSession.objects.create(
            uid='u-sub-refresh',
            nickname='老三',
            session_token='token-sub-refresh',
            token_expires_at=timezone.now() + timedelta(days=1),
            subscription_tier='free',
            subscription_display_name='免费版',
            limits={'max_classrooms': 3},
        )

        captured_paths = []

        def fake_request(method, url, body=None, headers=None, timeout=20):
            captured_paths.append(url)
            if url.endswith('/api/me/refresh-subscription'):
                return {
                    'ok': True,
                    'status': 'success',
                    'subscription': {
                        'tier': 'pro',
                        'display_name': '专业版',
                        'limits': {'max_classrooms': 99},
                    },
                }
            if url.endswith('/api/sync/status'):
                return {'ok': True, 'status': 'success', 'versions': {}}
            raise AssertionError(f'Unexpected URL: {url}')

        with patch('seats.cloud._request_json', side_effect=fake_request):
            payload = cloud_module.cloud_api_request(session, 'GET', '/api/sync/status')

        session.refresh_from_db()
        self.assertEqual(payload['status'], 'success')
        self.assertEqual(session.subscription_tier, 'pro')
        self.assertEqual(session.subscription_display_name, '专业版')
        self.assertEqual(session.limits.get('max_classrooms'), 99)
        self.assertEqual(len(captured_paths), 2)
        self.assertTrue(captured_paths[0].endswith('/api/me/refresh-subscription'))
        self.assertTrue(captured_paths[1].endswith('/api/sync/status'))


class BsceCloudSessionTests(TestCase):
    def test_bsce_cloud_list_reuses_browser_session_for_single_flow(self):
        from . import views

        class FakeHeaders:
            def __init__(self, set_cookie_headers=None):
                self.set_cookie_headers = list(set_cookie_headers or [])

            def get_all(self, name):
                if str(name).lower() == "set-cookie":
                    return list(self.set_cookie_headers)
                return []

        class FakeResponse:
            def __init__(self, payload, set_cookie_headers=None):
                self.payload = payload
                self.headers = FakeHeaders(set_cookie_headers)

            def __enter__(self):
                return self

            def __exit__(self, exc_type, exc, tb):
                return False

            def read(self):
                return json.dumps(self.payload).encode("utf-8")

        captured_requests = []

        def fake_urlopen(req, timeout=None, context=None):
            captured_requests.append(req)
            body = json.loads(req.data.decode("utf-8"))
            if body["action"] == "login":
                return FakeResponse({
                    "success": True,
                    "data": {
                        "username": "laosan",
                    },
                }, [
                    "sce_token=token-cookie; Domain=sce.jbyc.cc; Path=/; Max-Age=2592000; HttpOnly; SameSite=Lax",
                    "sce_username=laosan; Domain=sce.jbyc.cc; Path=/; Max-Age=2592000; HttpOnly; SameSite=Lax",
                ])
            if body["action"] == "get_settings":
                return FakeResponse({"success": True, "data": None})
            return FakeResponse({
                "success": True,
                "data": [
                    {
                        "fileId": "ws_1",
                        "metadata": {
                            "author": "laosan",
                            "name": "test",
                            "time": "2026-04-26T08:41:40+08:00",
                            "size": 4,
                        },
                    }
                ],
            })

        with patch("seats.views.secrets.token_hex", return_value="c" * 40):
            with patch("seats.views.secrets.choice", side_effect=[
                views.BSCE_CLOUD_BROWSER_PROFILES[0],
                views.BSCE_CLOUD_ACCEPT_LANGUAGES[0],
            ]):
                with patch("seats.views.urllib.request.urlopen", side_effect=fake_urlopen):
                    workspaces = views._bsce_cloud_list_workspaces("laosan", "secret")

        self.assertEqual(len(workspaces), 1)
        self.assertEqual([json.loads(req.data.decode("utf-8"))["action"] for req in captured_requests], [
            "login",
            "get_settings",
            "list",
        ])
        self.assertEqual(json.loads(captured_requests[1].data.decode("utf-8")).get("token"), "token-cookie")
        self.assertEqual(json.loads(captured_requests[2].data.decode("utf-8")).get("token"), "token-cookie")
        csrf_values = [json.loads(req.data.decode("utf-8")).get("_csrf") for req in captured_requests]
        self.assertEqual(csrf_values, ["c" * 40, "c" * 40, "c" * 40])

        user_agents = [req.get_header("User-agent") for req in captured_requests]
        cookies = [req.get_header("Cookie") for req in captured_requests]
        csrf_headers = [req.get_header("X-csrf-token") for req in captured_requests]
        self.assertEqual(len(set(user_agents)), 1)
        self.assertEqual(len(set(csrf_headers)), 1)
        self.assertIn("Edg/147.0.0.0", user_agents[0])
        self.assertIn("rth-uid=", cookies[0])
        self.assertIn(f"sce_csrf={'c' * 40}", cookies[0])
        self.assertNotIn("sce_token=token-cookie", cookies[0])
        self.assertIn("sce_token=token-cookie", cookies[1])
        self.assertIn("sce_username=laosan", cookies[1])
        self.assertEqual(cookies[1], cookies[2])
        self.assertEqual(captured_requests[0].get_header("Sec-ch-ua-platform"), '"macOS"')
        self.assertEqual(captured_requests[0].get_header("Accept-language"), views.BSCE_CLOUD_ACCEPT_LANGUAGES[0])

    def test_bsce_json_post_reports_http_403_html_body(self):
        html_body = (
            b"<!DOCTYPE html>\n"
            b"<html lang=\"zh-CN\"><head><meta charset=\"utf-8\"></head>"
            b"<body>Forbidden</body></html>"
        )
        error = urllib.error.HTTPError(
            url="https://sce.jbyc.cc/api/auth.php",
            code=403,
            msg="Forbidden",
            hdrs={},
            fp=BytesIO(html_body),
        )

        with patch("seats.views.urllib.request.urlopen", side_effect=error):
            with self.assertRaises(ValueError) as ctx:
                _bsce_json_post("https://sce.jbyc.cc/api/auth.php", {"action": "login"})

        message = str(ctx.exception)
        self.assertIn("云端请求失败：HTTP 403", message)
        self.assertIn("<!DOCTYPE html>", message)
        self.assertIn("zh-CN", message)


class BsceCloudImportErrorResponseTests(TestCase):
    def setUp(self):
        self.classroom = Classroom.objects.create(name="BSCE云导入测试班", rows=2, cols=2)

    def test_import_bsce_cloud_returns_json_when_cloud_rejects_request(self):
        url = reverse("import_bsce_cloud", args=[self.classroom.pk])

        with patch(
            "seats.views._bsce_cloud_list_workspaces",
            side_effect=ValueError("云端请求失败：HTTP 403 <!DOCTYPE html><html lang=\"zh-CN\">"),
        ):
            response = self.client.post(
                url,
                data=json.dumps({
                    "action": "list",
                    "username": "demo",
                    "password": "secret",
                }),
                content_type="application/json",
                HTTP_X_REQUESTED_WITH="XMLHttpRequest",
            )

        self.assertEqual(response.status_code, 400)
        payload = response.json()
        self.assertEqual(payload.get("status"), "error")
        self.assertIn("云端请求失败：HTTP 403", payload.get("message", ""))
        self.assertIn("<!DOCTYPE html>", payload.get("message", ""))


class FutureModeErrorHandlingTests(TestCase):
    def setUp(self):
        self.classroom = Classroom.objects.create(name="AI测试班", rows=2, cols=2)

    def test_ai_chat_maps_authentication_error_to_bad_request(self):
        url = reverse("ai_chat", args=[self.classroom.pk])
        request = httpx.Request("POST", "https://api.openai.com/v1/responses")
        response = httpx.Response(401, request=request)
        auth_error = openai.AuthenticationError(
            "Incorrect API key provided",
            response=response,
            body={"error": {"code": "invalid_api_key"}},
        )

        with patch("seats.views._run_future_mode", side_effect=auth_error):
            resp = self.client.post(
                url,
                data=json.dumps({"action": "message", "message": "你好"}),
                content_type="application/json",
            )

        self.assertEqual(resp.status_code, 400)
        payload = resp.json()
        self.assertEqual(payload.get("status"), "error")
        self.assertIn("鉴权失败", payload.get("message", ""))

    def test_ai_chat_maps_not_found_responses_error_to_bad_request(self):
        url = reverse("ai_chat", args=[self.classroom.pk])
        request = httpx.Request("POST", "https://example.com/v1/responses")
        response = httpx.Response(404, request=request)
        not_found_error = openai.NotFoundError(
            "No route for /responses",
            response=response,
            body={"error": {"message": "No route"}},
        )

        with patch("seats.views._run_future_mode", side_effect=not_found_error):
            resp = self.client.post(
                url,
                data=json.dumps({"action": "message", "message": "你好"}),
                content_type="application/json",
            )

        self.assertEqual(resp.status_code, 400)
        payload = resp.json()
        self.assertEqual(payload.get("status"), "error")
        self.assertIn("不支持 Responses API", payload.get("message", ""))

    def test_ai_chat_maps_not_implemented_text_error_to_bad_request(self):
        url = reverse("ai_chat", args=[self.classroom.pk])

        with patch("seats.views._run_future_mode", side_effect=Exception("status_code=500, not implemented")):
            resp = self.client.post(
                url,
                data=json.dumps({"action": "message", "message": "你好"}),
                content_type="application/json",
            )

        self.assertEqual(resp.status_code, 400)
        payload = resp.json()
        self.assertEqual(payload.get("status"), "error")
        self.assertIn("不支持 Responses API", payload.get("message", ""))

    def test_tool_approval_falls_back_when_call_id_mismatch(self):
        url = reverse("ai_chat", args=[self.classroom.pk])
        session = self.client.session
        session["future_mode_pending_tools"] = {
            "token_1": {
                "classroom_id": self.classroom.pk,
                "response_id": "resp_123",
                "function_calls": [
                    {
                        "call_id": "call_123",
                        "name": "get_classroom_overview",
                        "arguments": {},
                    }
                ],
            }
        }
        session.save()

        with patch(
            "seats.views._run_future_mode",
            side_effect=Exception("No tool call found for function call output with call_id call_123."),
        ):
            resp = self.client.post(
                url,
                data=json.dumps(
                    {
                        "action": "tool_approval",
                        "approval_token": "token_1",
                        "decisions": [{"call_id": "call_123", "approved": True}],
                    }
                ),
                content_type="application/json",
            )

        self.assertEqual(resp.status_code, 200)
        payload = resp.json()
        self.assertEqual(payload.get("status"), "success")
        self.assertIn("工具已执行", payload.get("reply", ""))
        self.assertIn("读取当前班级概览", payload.get("reply", ""))

    def test_tool_approval_returns_tool_error_to_ai_for_self_debugging(self):
        url = reverse("ai_chat", args=[self.classroom.pk])
        session = self.client.session
        session["future_mode_pending_tools"] = {
            "token_tool_err": {
                "classroom_id": self.classroom.pk,
                "response_id": "",
                "mode": "chat",
                "chat_messages": [{"role": "system", "content": "x"}],
                "function_calls": [
                    {
                        "call_id": "call_bad_query",
                        "name": "get_student_info",
                        "arguments": {"student_query": "不存在的学生"},
                    }
                ],
            }
        }
        session.save()

        with patch(
            "seats.views._run_future_mode",
            return_value={"status": "completed", "reply": "我已根据工具报错调整查询。", "tool_events": []},
        ) as mock_run:
            resp = self.client.post(
                url,
                data=json.dumps(
                    {
                        "action": "tool_approval",
                        "approval_token": "token_tool_err",
                        "decisions": [{"call_id": "call_bad_query", "approved": True}],
                    }
                ),
                content_type="application/json",
            )

        self.assertEqual(resp.status_code, 200)
        payload = resp.json()
        self.assertEqual(payload.get("status"), "success")

        tool_outputs = mock_run.call_args.kwargs.get("tool_outputs") or []
        self.assertEqual(len(tool_outputs), 1)
        tool_output_payload = json.loads(tool_outputs[0].get("output") or "{}")
        self.assertFalse(tool_output_payload.get("ok"))
        self.assertEqual(tool_output_payload.get("tool"), "get_student_info")
        self.assertIn("工具执行失败", tool_output_payload.get("message", ""))
        self.assertEqual((tool_output_payload.get("error") or {}).get("type"), "ValueError")


class FutureModeResponsesCompatibilityTests(TestCase):
    def test_create_future_mode_response_retries_with_content_parts(self):
        client = Mock()
        expected = object()
        client.responses.create.side_effect = [
            Exception("status_code=500, json: cannot unmarshal object into Go struct field ***.***.content of type []***.ResponsesOutputContent"),
            expected,
        ]

        result = _create_future_mode_response(
            client=client,
            model="gpt-4.1-mini",
            conversation=[{"role": "user", "content": "你好"}],
        )

        self.assertIs(result, expected)
        self.assertEqual(client.responses.create.call_count, 2)

        first_input = client.responses.create.call_args_list[0].kwargs["input"]
        second_input = client.responses.create.call_args_list[1].kwargs["input"]

        self.assertIsInstance(first_input[0]["content"], str)
        self.assertIsInstance(second_input[0]["content"], list)
        self.assertEqual(second_input[0]["content"][0]["type"], "input_text")


class FutureModeThinkingModeTests(TestCase):
    def setUp(self):
        self.classroom = Classroom.objects.create(name="思考模式测试班", rows=2, cols=2)

    def _build_chat_completion(self, text="ok"):
        assistant_message = SimpleNamespace(content=text, tool_calls=[])
        choice = SimpleNamespace(message=assistant_message)
        return SimpleNamespace(choices=[choice])

    def test_run_future_mode_chat_includes_extra_body_when_thinking_enabled(self):
        client = Mock()
        client.chat.completions.create.return_value = self._build_chat_completion("开启思考")

        result = _run_future_mode_chat(
            classroom=self.classroom,
            client=client,
            model="glm-5",
            conversation=[{"role": "user", "content": "你好"}],
            client_config={"thinking_mode": "enabled"},
        )

        self.assertEqual(result.get("status"), "completed")
        kwargs = client.chat.completions.create.call_args.kwargs
        self.assertEqual(kwargs.get("extra_body"), {"thinking": {"type": "enabled"}})

    def test_run_future_mode_chat_includes_extra_body_when_thinking_disabled(self):
        client = Mock()
        client.chat.completions.create.return_value = self._build_chat_completion("关闭思考")

        result = _run_future_mode_chat(
            classroom=self.classroom,
            client=client,
            model="glm-5",
            conversation=[{"role": "user", "content": "你好"}],
            client_config={"thinking_mode": "disabled"},
        )

        self.assertEqual(result.get("status"), "completed")
        kwargs = client.chat.completions.create.call_args.kwargs
        self.assertEqual(kwargs.get("extra_body"), {"thinking": {"type": "disabled"}})


class FutureModeConfigPersistenceTests(TestCase):
    def setUp(self):
        self.classroom = Classroom.objects.create(name="配置测试班", rows=2, cols=2)
        self.url = reverse("ai_chat", args=[self.classroom.pk])

    def test_config_save_and_get_roundtrip(self):
        save_resp = self.client.post(
            self.url,
            data=json.dumps(
                {
                    "action": "config_save",
                    "client_config": {
                        "api_key": "sk-db-123",
                        "base_url": "https://api.openai.com/v1",
                        "model": "gpt-4.1-mini",
                        "thinking_mode": "enabled",
                    },
                }
            ),
            content_type="application/json",
        )
        self.assertEqual(save_resp.status_code, 200)
        self.assertEqual(save_resp.json().get("status"), "success")

        db_config = FutureModeConfig.objects.get(classroom=self.classroom)
        self.assertEqual(db_config.api_key, "sk-db-123")
        self.assertEqual(db_config.base_url, "https://api.openai.com/v1")
        self.assertEqual(db_config.model, "gpt-4.1-mini")
        self.assertEqual(db_config.thinking_mode, "enabled")

        get_resp = self.client.post(
            self.url,
            data=json.dumps({"action": "config_get"}),
            content_type="application/json",
        )
        self.assertEqual(get_resp.status_code, 200)
        payload = get_resp.json()
        self.assertEqual(payload.get("status"), "success")
        self.assertEqual(payload.get("client_config", {}).get("api_key"), "sk-db-123")
        self.assertEqual(payload.get("client_config", {}).get("base_url"), "https://api.openai.com/v1")
        self.assertEqual(payload.get("client_config", {}).get("model"), "gpt-4.1-mini")
        self.assertEqual(payload.get("client_config", {}).get("thinking_mode"), "enabled")

    def test_message_uses_persisted_config_when_payload_empty(self):
        FutureModeConfig.objects.create(
            classroom=self.classroom,
            api_key="sk-db-x",
            base_url="https://api.openai.com/v1",
            model="gpt-4.1-mini",
        )

        with patch(
            "seats.views._run_future_mode",
            return_value={"status": "completed", "reply": "ok", "tool_events": []},
        ) as mock_run:
            resp = self.client.post(
                self.url,
                data=json.dumps({"action": "message", "message": "你好"}),
                content_type="application/json",
            )

        self.assertEqual(resp.status_code, 200)
        called_config = mock_run.call_args.kwargs.get("client_config", {})
        self.assertEqual(called_config.get("api_key"), "sk-db-x")
        self.assertEqual(called_config.get("base_url"), "https://api.openai.com/v1")
        self.assertEqual(called_config.get("model"), "gpt-4.1-mini")

    def test_payload_config_can_override_persisted_model(self):
        FutureModeConfig.objects.create(
            classroom=self.classroom,
            api_key="sk-db-y",
            base_url="https://api.openai.com/v1",
            model="gpt-4.1-mini",
            thinking_mode="",
        )

        with patch(
            "seats.views._run_future_mode",
            return_value={"status": "completed", "reply": "ok", "tool_events": []},
        ) as mock_run:
            resp = self.client.post(
                self.url,
                data=json.dumps(
                    {
                        "action": "message",
                        "message": "你好",
                        "client_config": {"model": "gpt-5-mini"},
                    }
                ),
                content_type="application/json",
            )

        self.assertEqual(resp.status_code, 200)
        called_config = mock_run.call_args.kwargs.get("client_config", {})
        self.assertEqual(called_config.get("api_key"), "sk-db-y")
        self.assertEqual(called_config.get("base_url"), "https://api.openai.com/v1")
        self.assertEqual(called_config.get("model"), "gpt-5-mini")

    def test_message_uses_persisted_thinking_mode_when_payload_empty(self):
        FutureModeConfig.objects.create(
            classroom=self.classroom,
            api_key="sk-db-z",
            base_url="https://api.openai.com/v1",
            model="gpt-4.1-mini",
            thinking_mode="enabled",
        )

        with patch(
            "seats.views._run_future_mode",
            return_value={"status": "completed", "reply": "ok", "tool_events": []},
        ) as mock_run:
            resp = self.client.post(
                self.url,
                data=json.dumps({"action": "message", "message": "你好"}),
                content_type="application/json",
            )

        self.assertEqual(resp.status_code, 200)
        called_config = mock_run.call_args.kwargs.get("client_config", {})
        self.assertEqual(called_config.get("thinking_mode"), "enabled")

    def test_message_branch_uses_auto_mode(self):
        with patch(
            "seats.views._run_future_mode",
            return_value={"status": "completed", "reply": "ok", "tool_events": []},
        ) as mock_run:
            resp = self.client.post(
                self.url,
                data=json.dumps({"action": "message", "message": "你好"}),
                content_type="application/json",
            )

        self.assertEqual(resp.status_code, 200)
        self.assertEqual(mock_run.call_args.kwargs.get("mode"), "auto")

    def test_tool_approval_uses_pending_chat_mode(self):
        session = self.client.session
        session["future_mode_pending_tools"] = {
            "token_chat": {
                "classroom_id": self.classroom.pk,
                "response_id": "",
                "mode": "chat",
                "chat_messages": [{"role": "system", "content": "x"}],
                "function_calls": [
                    {"call_id": "call_1", "name": "get_classroom_overview", "arguments": {}}
                ],
            }
        }
        session.save()

        with patch(
            "seats.views._run_future_mode",
            return_value={"status": "completed", "reply": "ok", "tool_events": []},
        ) as mock_run:
            resp = self.client.post(
                self.url,
                data=json.dumps(
                    {
                        "action": "tool_approval",
                        "approval_token": "token_chat",
                        "decisions": [{"call_id": "call_1", "approved": True}],
                    }
                ),
                content_type="application/json",
            )

        self.assertEqual(resp.status_code, 200)
        self.assertEqual(mock_run.call_args.kwargs.get("mode"), "chat")
        self.assertIsNone(mock_run.call_args.kwargs.get("previous_response_id"))
        self.assertEqual(mock_run.call_args.kwargs.get("chat_messages"), [{"role": "system", "content": "x"}])


class FutureModeDirectSwapTests(TestCase):
    def setUp(self):
        self.classroom = Classroom.objects.create(name="换座测试班", rows=1, cols=2)
        self.url = reverse("ai_chat", args=[self.classroom.pk])
        self.student_a = self.classroom.students.create(name="张三")
        self.student_b = self.classroom.students.create(name="李四")
        seat_1 = self.classroom.seats.get(row=1, col=1)
        seat_2 = self.classroom.seats.get(row=1, col=2)
        seat_1.student = self.student_a
        seat_1.save(update_fields=["student"])
        seat_2.student = self.student_b
        seat_2.save(update_fields=["student"])

    def test_message_detects_direct_swap_intent(self):
        resp = self.client.post(
            self.url,
            data=json.dumps({"action": "message", "message": "把张三和李四换位置"}),
            content_type="application/json",
        )
        self.assertEqual(resp.status_code, 200)
        payload = resp.json()
        self.assertEqual(payload.get("status"), "needs_approval")
        token = payload.get("approval_token")
        self.assertTrue(token)

        session_payload = self.client.session.get("future_mode_pending_tools", {}).get(token, {})
        self.assertEqual(session_payload.get("mode"), "direct")
        function_calls = session_payload.get("function_calls") or []
        self.assertEqual(len(function_calls), 1)
        self.assertEqual(function_calls[0].get("name"), "swap_students")

    def test_tool_approval_executes_direct_swap(self):
        session = self.client.session
        session["future_mode_pending_tools"] = {
            "token_swap": {
                "classroom_id": self.classroom.pk,
                "response_id": "",
                "mode": "direct",
                "chat_messages": [],
                "function_calls": [
                    {
                        "call_id": "call_swap_1",
                        "name": "swap_students",
                        "arguments": {"student_a": "张三", "student_b": "李四"},
                    }
                ],
            }
        }
        session.save()

        resp = self.client.post(
            self.url,
            data=json.dumps(
                {
                    "action": "tool_approval",
                    "approval_token": "token_swap",
                    "decisions": [{"call_id": "call_swap_1", "approved": True}],
                }
            ),
            content_type="application/json",
        )
        self.assertEqual(resp.status_code, 200)
        payload = resp.json()
        self.assertEqual(payload.get("status"), "success")

        self.student_a.refresh_from_db()
        self.student_b.refresh_from_db()
        self.assertEqual(self.student_a.assigned_seat.col, 2)
        self.assertEqual(self.student_b.assigned_seat.col, 1)


class ConstraintArrangeTests(TestCase):
    def test_must_together_does_not_assign_same_seat(self):
        classroom = Classroom.objects.create(name="T1", rows=2, cols=2)
        alice = classroom.students.create(name="Alice")
        bob = classroom.students.create(name="Bob")
        classroom.students.create(name="Carol")

        SeatConstraint.objects.create(
            classroom=classroom,
            constraint_type=SeatConstraint.ConstraintType.MUST_TOGETHER,
            student=alice,
            target_student=bob,
            distance=1,
        )

        students = list(classroom.students.all())
        seats = list(classroom.seats.select_related("student"))
        _arrange_standard(classroom, students, seats, "random")

        alice.refresh_from_db()
        bob.refresh_from_db()
        self.assertIsNotNone(alice.assigned_seat)
        self.assertIsNotNone(bob.assigned_seat)
        self.assertNotEqual(alice.assigned_seat.pk, bob.assigned_seat.pk)

        distance = abs(alice.assigned_seat.row - bob.assigned_seat.row) + abs(
            alice.assigned_seat.col - bob.assigned_seat.col
        )
        self.assertLessEqual(distance, 1)

    def test_special_internal_policy_keeps_working(self):
        classroom = Classroom.objects.create(name="T2", rows=2, cols=3)
        jqj = classroom.students.create(name="金千竣")
        hzh = classroom.students.create(name="胡哲豪")

        seat_jqj = classroom.seats.get(row=1, col=1)
        seat_hzh = classroom.seats.get(row=2, col=3)
        seat_jqj.student = jqj
        seat_jqj.save(update_fields=["student"])
        seat_hzh.student = hzh
        seat_hzh.save(update_fields=["student"])

        changed = _apply_internal_policy(classroom)
        self.assertTrue(changed)

        jqj.refresh_from_db()
        hzh.refresh_from_db()
        self.assertIsNotNone(jqj.assigned_seat)
        self.assertIsNotNone(hzh.assigned_seat)
        self.assertEqual(jqj.assigned_seat.row, hzh.assigned_seat.row)
        self.assertEqual(abs(jqj.assigned_seat.col - hzh.assigned_seat.col), 1)

    def test_group_mode_respects_must_seat_constraint(self):
        classroom = Classroom.objects.create(name="T3", rows=1, cols=4)
        g1 = SeatGroup.objects.create(classroom=classroom, name="G1", order=1)
        g2 = SeatGroup.objects.create(classroom=classroom, name="G2", order=2)

        for seat in classroom.seats.filter(cell_type=SeatCellType.SEAT):
            seat.group = g1 if seat.col <= 2 else g2
            seat.save(update_fields=["group"])

        s1 = classroom.students.create(name="A", score=100)
        classroom.students.create(name="B", score=0)

        SeatConstraint.objects.create(
            classroom=classroom,
            constraint_type=SeatConstraint.ConstraintType.MUST_SEAT,
            student=s1,
            row=1,
            col=2,
        )

        ok = _arrange_grouped(classroom, list(classroom.students.all()), "group_mentor")
        self.assertTrue(ok)

        s1.refresh_from_db()
        self.assertIsNotNone(s1.assigned_seat)
        self.assertEqual((s1.assigned_seat.row, s1.assigned_seat.col), (1, 2))


class GroupInteractionTests(TestCase):
    def test_apply_suggestion_disabled_type_returns_success(self):
        classroom = Classroom.objects.create(name="C0", rows=1, cols=2)
        url = reverse("apply_suggestion", args=[classroom.pk]) + "?type=jqj_hzh"
        response = self.client.post(url)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json().get("status"), "success")

    def test_apply_suggestion_swap_rejects_cross_class_students(self):
        c1 = Classroom.objects.create(name="C1", rows=1, cols=2)
        c2 = Classroom.objects.create(name="C2", rows=1, cols=2)
        s1 = c1.students.create(name="A")
        s2 = c2.students.create(name="B")
        c1.seats.get(row=1, col=1).student = s1
        c1.seats.get(row=1, col=1).save(update_fields=["student"])
        c2.seats.get(row=1, col=1).student = s2
        c2.seats.get(row=1, col=1).save(update_fields=["student"])

        url = reverse("apply_suggestion", args=[c1.pk]) + f"?type=swap_balance&s1={s1.pk}&s2={s2.pk}"
        response = self.client.post(url)
        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.json().get("status"), "error")

    def test_classroom_state_filters_internal_name_suggestions(self):
        classroom = Classroom.objects.create(name="C2A", rows=1, cols=2)
        g1 = SeatGroup.objects.create(classroom=classroom, name="G1", order=1)
        g2 = SeatGroup.objects.create(classroom=classroom, name="G2", order=2)

        seat1 = classroom.seats.get(row=1, col=1)
        seat2 = classroom.seats.get(row=1, col=2)
        seat1.group = g1
        seat2.group = g2
        seat1.save(update_fields=["group"])
        seat2.save(update_fields=["group"])

        s1 = classroom.students.create(name="金千竣", score=100)
        s2 = classroom.students.create(name="普通同学", score=10)
        seat1.student = s1
        seat2.student = s2
        seat1.save(update_fields=["student"])
        seat2.save(update_fields=["student"])

        state_url = reverse("classroom_state", args=[classroom.pk])
        response = self.client.get(state_url, HTTP_X_REQUESTED_WITH="XMLHttpRequest")
        self.assertEqual(response.status_code, 200)
        suggestions = response.json().get("suggestions", [])

        joined = []
        for item in suggestions:
            if isinstance(item, dict):
                joined.append(str(item.get("message") or ""))
                joined.append(str(item.get("action_url") or ""))
                joined.append(str(item.get("type") or ""))
            else:
                joined.append(str(item))
        text = " | ".join(joined)
        self.assertNotIn("金千竣", text)
        self.assertNotIn("胡哲豪", text)
        self.assertNotIn("jqj_hzh", text)

    def test_group_balance_does_not_suggest_internal_policy_students(self):
        classroom = Classroom.objects.create(name="C2B", rows=1, cols=4)
        g1 = SeatGroup.objects.create(classroom=classroom, name="G1", order=1)
        g2 = SeatGroup.objects.create(classroom=classroom, name="G2", order=2)

        seat1 = classroom.seats.get(row=1, col=1)
        seat2 = classroom.seats.get(row=1, col=2)
        seat3 = classroom.seats.get(row=1, col=3)
        seat4 = classroom.seats.get(row=1, col=4)
        seat1.group = g1
        seat2.group = g1
        seat3.group = g2
        seat4.group = g2
        seat1.save(update_fields=["group"])
        seat2.save(update_fields=["group"])
        seat3.save(update_fields=["group"])
        seat4.save(update_fields=["group"])

        s_internal = classroom.students.create(name="金千竣", score=100)
        s_high = classroom.students.create(name="高分甲", score=90)
        s_low1 = classroom.students.create(name="低分乙", score=5)
        s_low2 = classroom.students.create(name="低分丙", score=5)

        seat1.student = s_internal
        seat2.student = s_high
        seat3.student = s_low1
        seat4.student = s_low2
        seat1.save(update_fields=["student"])
        seat2.save(update_fields=["student"])
        seat3.save(update_fields=["student"])
        seat4.save(update_fields=["student"])

        state_url = reverse("classroom_state", args=[classroom.pk])
        response = self.client.get(state_url, HTTP_X_REQUESTED_WITH="XMLHttpRequest")
        self.assertEqual(response.status_code, 200)
        suggestions = response.json().get("suggestions", [])

        text = " | ".join(str(item) for item in suggestions)
        self.assertNotIn("金千竣", text)

    def test_set_podium_guards_updates_state_and_supports_partial_clear(self):
        classroom = Classroom.objects.create(name="护法班", rows=1, cols=2)
        left_student = classroom.students.create(name="左左", student_id="L001", score=91)
        right_student = classroom.students.create(name="右右", student_id="R001", score=87)

        left_seat = classroom.seats.get(row=1, col=1)
        right_seat = classroom.seats.get(row=1, col=2)
        left_seat.student = left_student
        left_seat.save(update_fields=["student"])
        right_seat.student = right_student
        right_seat.save(update_fields=["student"])

        response = self.client.post(
            reverse("set_podium_guards", args=[classroom.pk]),
            data=json.dumps({"left_student_id": left_student.pk, "right_student_id": right_student.pk}),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual((payload.get("podium_guards") or {}).get("left", {}).get("id"), left_student.pk)
        self.assertEqual((payload.get("podium_guards") or {}).get("right", {}).get("id"), right_student.pk)

        state_response = self.client.get(
            reverse("classroom_state", args=[classroom.pk]),
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )
        self.assertEqual(state_response.status_code, 200)
        state_payload = state_response.json()
        self.assertEqual((state_payload.get("podium_guards") or {}).get("left", {}).get("id"), left_student.pk)
        self.assertEqual((state_payload.get("podium_guards") or {}).get("right", {}).get("id"), right_student.pk)

        seat_rows = {f'{item["row"]}-{item["col"]}': item for item in state_payload.get("seats", [])}
        self.assertEqual((seat_rows.get("1-1") or {}).get("student", {}).get("podium_guardian_side"), "left")
        self.assertEqual((seat_rows.get("1-2") or {}).get("student", {}).get("podium_guardian_side"), "right")

        clear_response = self.client.post(
            reverse("set_podium_guards", args=[classroom.pk]),
            data=json.dumps({"right_student_id": ""}),
            content_type="application/json",
        )
        self.assertEqual(clear_response.status_code, 200)
        classroom.refresh_from_db()
        self.assertEqual(classroom.left_guardian_id, left_student.pk)
        self.assertIsNone(classroom.right_guardian_id)

    def test_classroom_state_auto_detects_podium_side_students(self):
        classroom = Classroom.objects.create(name="自动护法班", rows=1, cols=4)
        left_student = classroom.students.create(name="左左")
        right_student = classroom.students.create(name="右右")

        classroom.seats.filter(row=1, col=1).update(student=left_student)
        classroom.seats.filter(row=1, col=2).update(cell_type=SeatCellType.PODIUM, student=None, group=None)
        classroom.seats.filter(row=1, col=3).update(student=right_student)

        state_response = self.client.get(
            reverse("classroom_state", args=[classroom.pk]),
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )
        self.assertEqual(state_response.status_code, 200)
        payload = state_response.json()
        seat_rows = {f'{item["row"]}-{item["col"]}': item for item in payload.get("seats", [])}
        self.assertEqual((seat_rows.get("1-1") or {}).get("student", {}).get("podium_guardian_side"), "left")
        self.assertEqual((seat_rows.get("1-3") or {}).get("student", {}).get("podium_guardian_side"), "right")
        self.assertEqual((payload.get("podium_guards") or {}).get("left", {}).get("id"), left_student.pk)
        self.assertEqual((payload.get("podium_guards") or {}).get("right", {}).get("id"), right_student.pk)

    def test_classroom_detail_renders_unseated_podium_guardian_badges(self):
        classroom = Classroom.objects.create(name="未入座护法班", rows=1, cols=1)
        left_student = classroom.students.create(name="左左")
        right_student = classroom.students.create(name="右右")
        classroom.left_guardian = left_student
        classroom.right_guardian = right_student
        classroom.save(update_fields=["left_guardian", "right_guardian"])

        response = self.client.get(reverse("classroom_detail", args=[classroom.pk]))
        self.assertEqual(response.status_code, 200)
        html = response.content.decode("utf-8")
        self.assertIn("左左", html)
        self.assertIn("右右", html)
        self.assertIn('<span class="guardian-badge left-guard">左护法</span>', html)
        self.assertIn('<span class="guardian-badge right-guard">右护法</span>', html)

    def test_toggle_fixed_seat_creates_and_removes_must_seat_constraint(self):
        classroom = Classroom.objects.create(name="固定座位班", rows=1, cols=2)
        student = classroom.students.create(name="固定同学")
        classroom.seats.filter(row=1, col=1).update(student=student)

        enable_response = self.client.post(
            reverse("toggle_fixed_seat", args=[classroom.pk]),
            data=json.dumps({"row": 1, "col": 1, "enabled": True}),
            content_type="application/json",
        )
        self.assertEqual(enable_response.status_code, 200)
        self.assertTrue(enable_response.json().get("enabled"))

        constraint = classroom.constraints.get(
            student=student,
            constraint_type=SeatConstraint.ConstraintType.MUST_SEAT,
        )
        self.assertEqual((constraint.row, constraint.col), (1, 1))

        state_response = self.client.get(
            reverse("classroom_state", args=[classroom.pk]),
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )
        self.assertEqual(state_response.status_code, 200)
        seat_rows = {f'{item["row"]}-{item["col"]}': item for item in state_response.json().get("seats", [])}
        self.assertTrue((seat_rows.get("1-1") or {}).get("student", {}).get("is_fixed_seat"))

        disable_response = self.client.post(
            reverse("toggle_fixed_seat", args=[classroom.pk]),
            data=json.dumps({"row": 1, "col": 1, "enabled": False}),
            content_type="application/json",
        )
        self.assertEqual(disable_response.status_code, 200)
        self.assertFalse(disable_response.json().get("enabled"))
        self.assertFalse(
            classroom.constraints.filter(
                student=student,
                constraint_type=SeatConstraint.ConstraintType.MUST_SEAT,
            ).exists()
        )

    def test_set_podium_guards_rejects_same_student_for_both_sides(self):
        classroom = Classroom.objects.create(name="护法互斥班", rows=1, cols=1)
        student = classroom.students.create(name="同学甲")

        response = self.client.post(
            reverse("set_podium_guards", args=[classroom.pk]),
            data=json.dumps({"left_student_id": student.pk, "right_student_id": student.pk}),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.json().get("status"), "error")
        self.assertIn("不能设置为同一名学生", response.json().get("message", ""))

    def test_rename_group_duplicate_returns_error_in_ajax(self):
        classroom = Classroom.objects.create(name="C3", rows=1, cols=2)
        g1 = SeatGroup.objects.create(classroom=classroom, name="G1", order=1)
        SeatGroup.objects.create(classroom=classroom, name="G2", order=2)

        url = reverse("rename_group", args=[classroom.pk, g1.pk])
        response = self.client.post(
            url,
            {"name": "G2"},
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )
        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.json().get("status"), "error")

    def test_assign_group_clears_old_group_leader(self):
        classroom = Classroom.objects.create(name="C4", rows=1, cols=2)
        g1 = SeatGroup.objects.create(classroom=classroom, name="G1", order=1)
        g2 = SeatGroup.objects.create(classroom=classroom, name="G2", order=2)
        leader = classroom.students.create(name="Leader")
        seat = classroom.seats.get(row=1, col=1)
        seat.group = g1
        seat.student = leader
        seat.save(update_fields=["group", "student"])
        g1.leader = leader
        g1.save(update_fields=["leader"])

        url = reverse("assign_group", args=[classroom.pk])
        response = self.client.post(
            url,
            data=json.dumps({"row": 1, "col": 1, "group_id": g2.pk}),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        g1.refresh_from_db()
        self.assertIsNone(g1.leader_id)

    def test_move_student_auto_repairs_when_breaking_constraint(self):
        classroom = Classroom.objects.create(name="C5", rows=1, cols=2)
        s1 = classroom.students.create(name="A")
        s2 = classroom.students.create(name="B")
        seat1 = classroom.seats.get(row=1, col=1)
        seat2 = classroom.seats.get(row=1, col=2)
        seat1.student = s1
        seat1.save(update_fields=["student"])
        seat2.student = s2
        seat2.save(update_fields=["student"])

        SeatConstraint.objects.create(
            classroom=classroom,
            constraint_type=SeatConstraint.ConstraintType.MUST_SEAT,
            student=s1,
            row=1,
            col=1,
        )

        url = reverse("move_student", args=[classroom.pk])
        response = self.client.post(
            url,
            data=json.dumps({"student_id": s1.pk, "row": 1, "col": 2}),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json().get("status"), "success")
        seat1.refresh_from_db()
        seat2.refresh_from_db()
        self.assertEqual(seat1.student_id, s1.pk)
        self.assertEqual(seat2.student_id, s2.pk)

    def test_first_move_of_special_student_keeps_target_position(self):
        classroom = Classroom.objects.create(name="C5S", rows=2, cols=3)
        jqj = classroom.students.create(name="金千竣")
        hzh = classroom.students.create(name="胡哲豪")

        seat_jqj = classroom.seats.get(row=1, col=1)
        seat_hzh = classroom.seats.get(row=1, col=2)
        seat_jqj.student = jqj
        seat_jqj.save(update_fields=["student"])
        seat_hzh.student = hzh
        seat_hzh.save(update_fields=["student"])

        url = reverse("move_student", args=[classroom.pk])
        response = self.client.post(
            url,
            data=json.dumps({"student_id": jqj.pk, "row": 2, "col": 3}),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json().get("status"), "success")

        jqj.refresh_from_db()
        hzh.refresh_from_db()
        self.assertEqual((jqj.assigned_seat.row, jqj.assigned_seat.col), (2, 3))
        distance = abs(jqj.assigned_seat.row - hzh.assigned_seat.row) + abs(
            jqj.assigned_seat.col - hzh.assigned_seat.col
        )
        self.assertEqual(distance, 1)

    def test_move_students_batch_moves_multiple_students(self):
        classroom = Classroom.objects.create(name="C5B", rows=2, cols=3)
        s1 = classroom.students.create(name="A")
        s2 = classroom.students.create(name="B")
        seat_a = classroom.seats.get(row=1, col=1)
        seat_b = classroom.seats.get(row=1, col=2)
        seat_a.student = s1
        seat_a.save(update_fields=["student"])
        seat_b.student = s2
        seat_b.save(update_fields=["student"])

        url = reverse("move_students_batch", args=[classroom.pk])
        response = self.client.post(
            url,
            data=json.dumps(
                {
                    "moves": [
                        {"student_id": s1.pk, "row": 2, "col": 1},
                        {"student_id": s2.pk, "row": 2, "col": 2},
                    ]
                }
            ),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json().get("status"), "success")

        s1.refresh_from_db()
        s2.refresh_from_db()
        self.assertEqual((s1.assigned_seat.row, s1.assigned_seat.col), (2, 1))
        self.assertEqual((s2.assigned_seat.row, s2.assigned_seat.col), (2, 2))

    def test_move_students_batch_supports_clear_then_assign(self):
        classroom = Classroom.objects.create(name="C5B-QuickSwap", rows=1, cols=2)
        s1 = classroom.students.create(name="A")
        s2 = classroom.students.create(name="B")
        seat_a = classroom.seats.get(row=1, col=1)
        seat_a.student = s1
        seat_a.save(update_fields=["student"])

        url = reverse("move_students_batch", args=[classroom.pk])
        response = self.client.post(
            url,
            data=json.dumps(
                {
                    "moves": [
                        {"student_id": s1.pk, "row": None, "col": None},
                        {"student_id": s2.pk, "row": 1, "col": 1},
                    ]
                }
            ),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json().get("status"), "success")

        seat_a.refresh_from_db()
        self.assertEqual(seat_a.student_id, s2.pk)
        self.assertFalse(classroom.seats.filter(student=s1).exists())

    def test_move_students_batch_rejects_duplicate_target(self):
        classroom = Classroom.objects.create(name="C5C", rows=2, cols=2)
        s1 = classroom.students.create(name="A")
        s2 = classroom.students.create(name="B")
        seat_a = classroom.seats.get(row=1, col=1)
        seat_b = classroom.seats.get(row=1, col=2)
        seat_a.student = s1
        seat_a.save(update_fields=["student"])
        seat_b.student = s2
        seat_b.save(update_fields=["student"])

        url = reverse("move_students_batch", args=[classroom.pk])
        response = self.client.post(
            url,
            data=json.dumps(
                {
                    "moves": [
                        {"student_id": s1.pk, "row": 2, "col": 1},
                        {"student_id": s2.pk, "row": 2, "col": 1},
                    ]
                }
            ),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.json().get("status"), "error")

    def test_move_students_batch_supports_undo_redo(self):
        classroom = Classroom.objects.create(name="C5D", rows=2, cols=2)
        s1 = classroom.students.create(name="A")
        s2 = classroom.students.create(name="B")
        seat_a = classroom.seats.get(row=1, col=1)
        seat_b = classroom.seats.get(row=1, col=2)
        seat_a.student = s1
        seat_a.save(update_fields=["student"])
        seat_b.student = s2
        seat_b.save(update_fields=["student"])

        move_url = reverse("move_students_batch", args=[classroom.pk])
        self.client.post(
            move_url,
            data=json.dumps(
                {
                    "moves": [
                        {"student_id": s1.pk, "row": 2, "col": 1},
                        {"student_id": s2.pk, "row": 2, "col": 2},
                    ]
                }
            ),
            content_type="application/json",
        )

        undo_url = reverse("undo_action", args=[classroom.pk])
        redo_url = reverse("redo_action", args=[classroom.pk])

        undo_resp = self.client.post(undo_url)
        self.assertEqual(undo_resp.status_code, 200)
        s1.refresh_from_db()
        s2.refresh_from_db()
        self.assertEqual((s1.assigned_seat.row, s1.assigned_seat.col), (1, 1))
        self.assertEqual((s2.assigned_seat.row, s2.assigned_seat.col), (1, 2))

        redo_resp = self.client.post(redo_url)
        self.assertEqual(redo_resp.status_code, 200)
        s1.refresh_from_db()
        s2.refresh_from_db()
        self.assertEqual((s1.assigned_seat.row, s1.assigned_seat.col), (2, 1))
        self.assertEqual((s2.assigned_seat.row, s2.assigned_seat.col), (2, 2))

    def test_move_student_supports_group_follow_mode(self):
        classroom = Classroom.objects.create(name="C5GF", rows=1, cols=2)
        g1 = SeatGroup.objects.create(classroom=classroom, name="第一组", order=1)
        g2 = SeatGroup.objects.create(classroom=classroom, name="第二组", order=2)
        s1 = classroom.students.create(name="甲")
        s2 = classroom.students.create(name="乙")
        seat_a = classroom.seats.get(row=1, col=1)
        seat_b = classroom.seats.get(row=1, col=2)
        seat_a.student = s1
        seat_a.group = g1
        seat_a.save(update_fields=["student", "group"])
        seat_b.student = s2
        seat_b.group = g2
        seat_b.save(update_fields=["student", "group"])

        response = self.client.post(
            reverse("move_student", args=[classroom.pk]),
            data=json.dumps({"student_id": s1.pk, "row": 1, "col": 2, "group_move_mode": "follow"}),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json().get("status"), "success")
        seat_a.refresh_from_db()
        seat_b.refresh_from_db()
        self.assertEqual(seat_a.student_id, s2.pk)
        self.assertEqual(seat_a.group_id, g2.pk)
        self.assertEqual(seat_b.student_id, s1.pk)
        self.assertEqual(seat_b.group_id, g1.pk)

    def test_move_students_batch_supports_group_follow_mode(self):
        classroom = Classroom.objects.create(name="C5GB", rows=2, cols=2)
        g1 = SeatGroup.objects.create(classroom=classroom, name="第一组", order=1)
        g2 = SeatGroup.objects.create(classroom=classroom, name="第二组", order=2)
        s1 = classroom.students.create(name="甲")
        s2 = classroom.students.create(name="乙")
        seat_a = classroom.seats.get(row=1, col=1)
        seat_b = classroom.seats.get(row=1, col=2)
        seat_c = classroom.seats.get(row=2, col=1)
        seat_d = classroom.seats.get(row=2, col=2)
        seat_a.student = s1
        seat_a.group = g1
        seat_a.save(update_fields=["student", "group"])
        seat_b.student = s2
        seat_b.group = g2
        seat_b.save(update_fields=["student", "group"])

        response = self.client.post(
            reverse("move_students_batch", args=[classroom.pk]),
            data=json.dumps(
                {
                    "group_move_mode": "follow",
                    "moves": [
                        {"student_id": s1.pk, "row": 2, "col": 1},
                        {"student_id": s2.pk, "row": 2, "col": 2},
                    ],
                }
            ),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json().get("status"), "success")
        seat_a.refresh_from_db()
        seat_b.refresh_from_db()
        seat_c.refresh_from_db()
        seat_d.refresh_from_db()
        self.assertIsNone(seat_a.student_id)
        self.assertIsNone(seat_a.group_id)
        self.assertIsNone(seat_b.student_id)
        self.assertIsNone(seat_b.group_id)
        self.assertEqual(seat_c.student_id, s1.pk)
        self.assertEqual(seat_c.group_id, g1.pk)
        self.assertEqual(seat_d.student_id, s2.pk)
        self.assertEqual(seat_d.group_id, g2.pk)

    def test_swap_suggestion_auto_repairs_when_breaking_constraint(self):
        classroom = Classroom.objects.create(name="C6", rows=1, cols=2)
        s1 = classroom.students.create(name="A")
        s2 = classroom.students.create(name="B")
        seat1 = classroom.seats.get(row=1, col=1)
        seat2 = classroom.seats.get(row=1, col=2)
        seat1.student = s1
        seat1.save(update_fields=["student"])
        seat2.student = s2
        seat2.save(update_fields=["student"])

        SeatConstraint.objects.create(
            classroom=classroom,
            constraint_type=SeatConstraint.ConstraintType.MUST_SEAT,
            student=s1,
            row=1,
            col=1,
        )

        url = reverse("apply_suggestion", args=[classroom.pk]) + f"?type=swap_balance&s1={s1.pk}&s2={s2.pk}"
        response = self.client.post(url)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json().get("status"), "success")
        seat1.refresh_from_db()
        seat2.refresh_from_db()
        self.assertEqual(seat1.student_id, s1.pk)
        self.assertEqual(seat2.student_id, s2.pk)

    def test_auto_group_nearby_uses_shape_profile(self):
        classroom = Classroom.objects.create(name="C7", rows=6, cols=6)
        ref_group = SeatGroup.objects.create(classroom=classroom, name="1", order=1)

        ref_coords = [(5, 5), (5, 6), (6, 5), (6, 6)]
        for idx, (r, c) in enumerate(ref_coords, start=1):
            stu = classroom.students.create(name=f"Ref{idx}")
            seat = classroom.seats.get(row=r, col=c)
            seat.student = stu
            seat.group = ref_group
            seat.save(update_fields=["student", "group"])

        line_coords = [(1, 1), (1, 2), (1, 3), (1, 4)]
        block_coords = [(2, 1), (2, 2), (3, 1), (3, 2)]
        target_coords = line_coords + block_coords
        for idx, (r, c) in enumerate(target_coords, start=1):
            stu = classroom.students.create(name=f"T{idx}")
            seat = classroom.seats.get(row=r, col=c)
            seat.student = stu
            seat.group = None
            seat.save(update_fields=["student", "group"])

        url = reverse("auto_group_from_reference", args=[classroom.pk])
        response = self.client.post(
            url,
            data=json.dumps(
                {
                    "reference_group_id": ref_group.pk,
                    "remainder_strategy": "merge_prev",
                    "auto_detect_group_style": True,
                }
            ),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload.get("status"), "success")
        self.assertEqual(payload.get("group_style"), "nearby")
        self.assertEqual(payload.get("group_shape"), "block_2x2")

        created_groups = payload.get("created_groups") or []
        self.assertGreaterEqual(len(created_groups), 2)
        first_group_id = created_groups[0]["id"]
        first_group_coords = set(
            classroom.seats.filter(
                group_id=first_group_id,
                row__in=[1, 2, 3],
                col__in=[1, 2, 3, 4],
            ).values_list("row", "col")
        )
        self.assertEqual(len(first_group_coords), 4)
        min_row = min(r for r, _ in first_group_coords)
        min_col = min(c for _, c in first_group_coords)
        normalized = {(r - min_row, c - min_col) for r, c in first_group_coords}
        self.assertEqual(normalized, {(0, 0), (0, 1), (1, 0), (1, 1)})

    def test_auto_group_nearby_tiles_three_rows_and_puts_remainder_to_one_group(self):
        classroom = Classroom.objects.create(name="C8", rows=7, cols=4)
        ref_group = SeatGroup.objects.create(classroom=classroom, name="1", order=1)

        ref_coords = [(5, 1), (5, 2), (6, 1), (6, 2), (7, 1), (7, 2)]
        for idx, (r, c) in enumerate(ref_coords, start=1):
            stu = classroom.students.create(name=f"RefG{idx}")
            seat = classroom.seats.get(row=r, col=c)
            seat.student = stu
            seat.group = ref_group
            seat.save(update_fields=["student", "group"])

        target_coords = [(r, c) for r in range(1, 5) for c in range(1, 5)]
        for idx, (r, c) in enumerate(target_coords, start=1):
            stu = classroom.students.create(name=f"S{idx}")
            seat = classroom.seats.get(row=r, col=c)
            seat.student = stu
            seat.group = None
            seat.save(update_fields=["student", "group"])

        url = reverse("auto_group_from_reference", args=[classroom.pk])
        response = self.client.post(
            url,
            data=json.dumps(
                {
                    "reference_group_id": ref_group.pk,
                    "remainder_strategy": "new_group",
                    "auto_detect_group_style": True,
                }
            ),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload.get("status"), "success")
        self.assertEqual(payload.get("group_style"), "nearby")
        self.assertEqual(payload.get("group_shape"), "block_3x2")

        created_groups = payload.get("created_groups") or []
        self.assertGreaterEqual(len(created_groups), 3)
        group_ids = [g["id"] for g in created_groups[:3]]

        counts = []
        group_coords = {}
        for gid in group_ids:
            coords = set(
                classroom.seats.filter(group_id=gid).values_list("row", "col")
            )
            group_coords[gid] = coords
            counts.append(len(coords))
        self.assertEqual(sorted(counts), [4, 6, 6])

        remainder_groups = [gid for gid in group_ids if len(group_coords[gid]) == 4]
        self.assertEqual(len(remainder_groups), 1)
        remainder_coords = group_coords[remainder_groups[0]]
        self.assertEqual(remainder_coords, {(4, 1), (4, 2), (4, 3), (4, 4)})

    def test_auto_group_horizontal_ignores_group_size_and_groups_by_row(self):
        classroom = Classroom.objects.create(name="C9", rows=4, cols=4)
        ref_group = SeatGroup.objects.create(classroom=classroom, name="1", order=1)

        ref_coords = [(4, 1), (4, 2)]
        for idx, (r, c) in enumerate(ref_coords, start=1):
            stu = classroom.students.create(name=f"RefH{idx}")
            seat = classroom.seats.get(row=r, col=c)
            seat.student = stu
            seat.group = ref_group
            seat.save(update_fields=["student", "group"])

        target_coords = [(r, c) for r in [1, 2] for c in [1, 2, 3, 4]]
        for idx, (r, c) in enumerate(target_coords, start=1):
            stu = classroom.students.create(name=f"H{idx}")
            seat = classroom.seats.get(row=r, col=c)
            seat.student = stu
            seat.group = None
            seat.save(update_fields=["student", "group"])

        url = reverse("auto_group_from_reference", args=[classroom.pk])
        response = self.client.post(
            url,
            data=json.dumps(
                {
                    "reference_group_id": ref_group.pk,
                    "remainder_strategy": "skip",
                    "auto_detect_group_style": True,
                }
            ),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload.get("status"), "success")
        self.assertEqual(payload.get("group_style"), "horizontal")
        self.assertTrue(payload.get("linear_grouping"))

        created_groups = payload.get("created_groups") or []
        self.assertEqual(len(created_groups), 2)

        group_rows = {}
        for g in created_groups:
            coords = set(classroom.seats.filter(group_id=g["id"]).values_list("row", "col"))
            self.assertEqual(len(coords), 4)
            rows = {r for r, _ in coords}
            self.assertEqual(len(rows), 1)
            group_rows[g["id"]] = rows.pop()
        self.assertEqual(set(group_rows.values()), {1, 2})

    def test_auto_group_vertical_ignores_group_size_and_groups_by_col(self):
        classroom = Classroom.objects.create(name="C10", rows=4, cols=4)
        ref_group = SeatGroup.objects.create(classroom=classroom, name="1", order=1)

        ref_coords = [(1, 4), (2, 4)]
        for idx, (r, c) in enumerate(ref_coords, start=1):
            stu = classroom.students.create(name=f"RefV{idx}")
            seat = classroom.seats.get(row=r, col=c)
            seat.student = stu
            seat.group = ref_group
            seat.save(update_fields=["student", "group"])

        target_coords = [(r, c) for c in [1, 2] for r in [1, 2, 3, 4]]
        for idx, (r, c) in enumerate(target_coords, start=1):
            stu = classroom.students.create(name=f"V{idx}")
            seat = classroom.seats.get(row=r, col=c)
            seat.student = stu
            seat.group = None
            seat.save(update_fields=["student", "group"])

        url = reverse("auto_group_from_reference", args=[classroom.pk])
        response = self.client.post(
            url,
            data=json.dumps(
                {
                    "reference_group_id": ref_group.pk,
                    "remainder_strategy": "skip",
                    "auto_detect_group_style": True,
                }
            ),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload.get("status"), "success")
        self.assertEqual(payload.get("group_style"), "vertical")
        self.assertTrue(payload.get("linear_grouping"))

        created_groups = payload.get("created_groups") or []
        self.assertEqual(len(created_groups), 2)

        group_cols = {}
        for g in created_groups:
            coords = set(classroom.seats.filter(group_id=g["id"]).values_list("row", "col"))
            self.assertEqual(len(coords), 4)
            cols = {c for _, c in coords}
            self.assertEqual(len(cols), 1)
            group_cols[g["id"]] = cols.pop()
        self.assertEqual(set(group_cols.values()), {1, 2})


class ClassroomHistoryTests(TestCase):
    def test_history_is_persisted_in_database_across_clients(self):
        classroom = Classroom.objects.create(name="历史持久化班", rows=1, cols=2)
        student = classroom.students.create(name="张三")
        seat1 = classroom.seats.get(row=1, col=1)
        seat1.student = student
        seat1.save(update_fields=["student"])

        move_url = reverse("move_student", args=[classroom.pk])
        response = self.client.post(
            move_url,
            data=json.dumps({"student_id": student.pk, "row": 1, "col": 2}),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(ClassroomHistoryEntry.objects.filter(classroom=classroom).count(), 1)

        another_client = Client()
        undo_response = another_client.post(reverse("undo_action", args=[classroom.pk]))
        self.assertEqual(undo_response.status_code, 200)
        student.refresh_from_db()
        self.assertEqual((student.assigned_seat.row, student.assigned_seat.col), (1, 1))

        redo_response = another_client.post(reverse("redo_action", args=[classroom.pk]))
        self.assertEqual(redo_response.status_code, 200)
        student.refresh_from_db()
        self.assertEqual((student.assigned_seat.row, student.assigned_seat.col), (1, 2))

    def test_merge_groups_supports_undo_and_redo(self):
        classroom = Classroom.objects.create(name="合组回滚班", rows=1, cols=2)
        g1 = SeatGroup.objects.create(classroom=classroom, name="A组", order=1)
        g2 = SeatGroup.objects.create(classroom=classroom, name="B组", order=2)
        s1 = classroom.students.create(name="甲")
        s2 = classroom.students.create(name="乙")

        seat1 = classroom.seats.get(row=1, col=1)
        seat2 = classroom.seats.get(row=1, col=2)
        seat1.student = s1
        seat1.group = g1
        seat1.save(update_fields=["student", "group"])
        seat2.student = s2
        seat2.group = g2
        seat2.save(update_fields=["student", "group"])
        g2.leader = s2
        g2.save(update_fields=["leader"])

        merge_url = reverse("merge_groups", args=[classroom.pk])
        response = self.client.post(
            merge_url,
            data=json.dumps({"target_group_id": g1.pk, "source_group_ids": [g2.pk]}),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        self.assertFalse(SeatGroup.objects.filter(pk=g2.pk).exists())

        undo_response = self.client.post(reverse("undo_action", args=[classroom.pk]))
        self.assertEqual(undo_response.status_code, 200)
        g2.refresh_from_db()
        seat2.refresh_from_db()
        self.assertEqual(seat2.group_id, g2.pk)
        self.assertEqual(g2.leader_id, s2.pk)

        redo_response = self.client.post(reverse("redo_action", args=[classroom.pk]))
        self.assertEqual(redo_response.status_code, 200)
        self.assertFalse(SeatGroup.objects.filter(pk=g2.pk).exists())
        seat2.refresh_from_db()
        self.assertEqual(seat2.group_id, g1.pk)

    def test_save_layout_snapshot_supports_undo_and_redo(self):
        classroom = Classroom.objects.create(name="快照回滚班", rows=1, cols=1)

        save_url = reverse("save_layout_snapshot", args=[classroom.pk])
        response = self.client.post(save_url, {"snapshot_name": "期中布局"})
        self.assertEqual(response.status_code, 302)
        snapshot = LayoutSnapshot.objects.get(classroom=classroom, name="期中布局")

        undo_response = self.client.post(reverse("undo_action", args=[classroom.pk]))
        self.assertEqual(undo_response.status_code, 200)
        self.assertFalse(LayoutSnapshot.objects.filter(pk=snapshot.pk).exists())

        redo_response = self.client.post(reverse("redo_action", args=[classroom.pk]))
        self.assertEqual(redo_response.status_code, 200)
        self.assertTrue(LayoutSnapshot.objects.filter(pk=snapshot.pk, name="期中布局").exists())

    def test_set_group_leader_supports_undo_and_redo(self):
        classroom = Classroom.objects.create(name="组长回滚班", rows=1, cols=1)
        group = SeatGroup.objects.create(classroom=classroom, name="第一组", order=1)
        student = classroom.students.create(name="班长")
        seat = classroom.seats.get(row=1, col=1)
        seat.student = student
        seat.group = group
        seat.save(update_fields=["student", "group"])

        url = reverse("set_group_leader", args=[classroom.pk])
        response = self.client.post(
            url,
            data=json.dumps({"student_id": student.pk}),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        group.refresh_from_db()
        self.assertEqual(group.leader_id, student.pk)

        undo_response = self.client.post(reverse("undo_action", args=[classroom.pk]))
        self.assertEqual(undo_response.status_code, 200)
        group.refresh_from_db()
        self.assertIsNone(group.leader_id)

        redo_response = self.client.post(reverse("redo_action", args=[classroom.pk]))
        self.assertEqual(redo_response.status_code, 200)
        group.refresh_from_db()
        self.assertEqual(group.leader_id, student.pk)

    def test_set_podium_guards_supports_undo_and_redo(self):
        classroom = Classroom.objects.create(name="护法回滚班", rows=1, cols=2)
        left_student = classroom.students.create(name="左护法")
        right_student = classroom.students.create(name="右护法")

        response = self.client.post(
            reverse("set_podium_guards", args=[classroom.pk]),
            data=json.dumps({"left_student_id": left_student.pk, "right_student_id": right_student.pk}),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        classroom.refresh_from_db()
        self.assertEqual(classroom.left_guardian_id, left_student.pk)
        self.assertEqual(classroom.right_guardian_id, right_student.pk)

        undo_response = self.client.post(reverse("undo_action", args=[classroom.pk]))
        self.assertEqual(undo_response.status_code, 200)
        classroom.refresh_from_db()
        self.assertIsNone(classroom.left_guardian_id)
        self.assertIsNone(classroom.right_guardian_id)

        redo_response = self.client.post(reverse("redo_action", args=[classroom.pk]))
        self.assertEqual(redo_response.status_code, 200)
        classroom.refresh_from_db()
        self.assertEqual(classroom.left_guardian_id, left_student.pk)
        self.assertEqual(classroom.right_guardian_id, right_student.pk)

    def test_history_only_keeps_latest_thousand_entries(self):
        classroom = Classroom.objects.create(name="历史上限班", rows=1, cols=1)
        url = reverse("rename_classroom", args=[classroom.pk])

        for index in range(1005):
            response = self.client.post(
                url,
                data=json.dumps({"name": f"历史班{index}"}),
                content_type="application/json",
                HTTP_X_REQUESTED_WITH="XMLHttpRequest",
            )
            self.assertEqual(response.status_code, 200)

        entries = list(ClassroomHistoryEntry.objects.filter(classroom=classroom).order_by("pk"))
        self.assertEqual(len(entries), 1000)
        self.assertEqual(entries[0].payload.get("type"), "rename_classroom")
        self.assertEqual(entries[0].payload.get("name"), "历史班5")
        self.assertTrue(all(entry.is_applied for entry in entries))

    def test_exported_seats_file_restores_full_bundle_and_undo_history(self):
        source = Classroom.objects.create(name="完整导出班", rows=1, cols=2)
        student = source.students.create(name="张三", student_id="001", gender="M", score=88)
        assistant_student = source.students.create(name="李四", student_id="002", gender="F", score=80)
        group = SeatGroup.objects.create(classroom=source, name="第一组", order=1)

        source_seat = source.seats.get(row=1, col=1)
        source_seat.student = student
        source_seat.group = group
        source_seat.save(update_fields=["student", "group"])
        target_seat = source.seats.get(row=1, col=2)
        target_seat.student = assistant_student
        target_seat.group = group
        target_seat.save(update_fields=["student", "group"])
        group.leader = student
        group.save(update_fields=["leader"])
        source.left_guardian = student
        source.right_guardian = assistant_student
        source.save(update_fields=["left_guardian", "right_guardian"])

        SeatConstraint.objects.create(
            classroom=source,
            constraint_type=SeatConstraint.ConstraintType.MUST_ROW,
            student=student,
            row=1,
            note="原位",
        )
        FutureModeConfig.objects.create(
            classroom=source,
            api_key="sk-test",
            base_url="https://example.com/v1",
            model="gpt-test",
            thinking_mode="enabled",
        )
        conversation = AIConversation.objects.create(
            classroom=source,
            session_key="source-session",
            title="排座记录",
            last_mode="chat",
            last_response_id="resp_1",
        )
        AIConversationMessage.objects.create(
            conversation=conversation,
            role=AIConversationMessage.MessageRole.USER,
            content="你好",
            payload={"cards": [{"type": "student_detail", "title": "测试卡片"}]},
        )
        AIConversationMessage.objects.create(
            conversation=conversation,
            role=AIConversationMessage.MessageRole.ASSISTANT,
            content="好的",
            payload={},
        )

        save_response = self.client.post(
            reverse("save_layout_snapshot", args=[source.pk]),
            {"snapshot_name": "开局"},
        )
        self.assertEqual(save_response.status_code, 302)

        move_response = self.client.post(
            reverse("move_student", args=[source.pk]),
            data=json.dumps({"student_id": student.pk, "row": 1, "col": 2}),
            content_type="application/json",
        )
        self.assertEqual(move_response.status_code, 200)

        export_response = self.client.get(reverse("export_seats_file", args=[source.pk]))
        self.assertEqual(export_response.status_code, 200)
        export_payload = json.loads(export_response.content.decode("utf-8"))
        self.assertEqual(export_payload.get("meta", {}).get("schema"), "full")
        self.assertIn("current_state", export_payload)
        self.assertIn("history", export_payload)
        self.assertIn("future_mode_config", export_payload)
        self.assertIn("ai_conversations", export_payload)

        target = Classroom.objects.create(name="导入目标班", rows=1, cols=1)
        upload = SimpleUploadedFile(
            "demo.seats",
            json.dumps(export_payload, ensure_ascii=False).encode("utf-8"),
            content_type="application/octet-stream",
        )
        import_response = self.client.post(
            reverse("import_seats_file", args=[target.pk]),
            {"seats_file": upload},
        )
        self.assertEqual(import_response.status_code, 302)

        target.refresh_from_db()
        self.assertEqual(target.name, "完整导出班")

        imported_student = Student.objects.get(classroom=target, student_id="001")
        imported_assistant_student = Student.objects.get(classroom=target, student_id="002")
        self.assertEqual((imported_student.assigned_seat.row, imported_student.assigned_seat.col), (1, 2))

        imported_group = SeatGroup.objects.get(classroom=target, name="第一组")
        self.assertEqual(imported_group.leader_id, imported_student.pk)
        self.assertEqual(target.left_guardian_id, imported_student.pk)
        self.assertEqual(target.right_guardian_id, imported_assistant_student.pk)

        imported_snapshot = LayoutSnapshot.objects.get(classroom=target, name="开局")
        imported_config = FutureModeConfig.objects.get(classroom=target)
        self.assertEqual(imported_config.model, "gpt-test")
        self.assertEqual(imported_config.base_url, "https://example.com/v1")

        imported_conversation = AIConversation.objects.get(classroom=target)
        self.assertEqual(imported_conversation.title, "排座记录")
        self.assertEqual(imported_conversation.messages.count(), 2)

        self.assertEqual(
            ClassroomHistoryEntry.objects.filter(classroom=target).count(),
            ClassroomHistoryEntry.objects.filter(classroom=source).count(),
        )

        undo_response = self.client.post(reverse("undo_action", args=[target.pk]))
        self.assertEqual(undo_response.status_code, 200)
        imported_student.refresh_from_db()
        self.assertEqual((imported_student.assigned_seat.row, imported_student.assigned_seat.col), (1, 1))

        redo_response = self.client.post(reverse("redo_action", args=[target.pk]))
        self.assertEqual(redo_response.status_code, 200)
        imported_student.refresh_from_db()
        self.assertEqual((imported_student.assigned_seat.row, imported_student.assigned_seat.col), (1, 2))

        load_response = self.client.get(reverse("load_layout_snapshot", args=[target.pk, imported_snapshot.pk]))
        self.assertEqual(load_response.status_code, 302)
        imported_student.refresh_from_db()
        self.assertEqual((imported_student.assigned_seat.row, imported_student.assigned_seat.col), (1, 1))

    def test_legacy_seats_file_import_keeps_old_behavior(self):
        classroom = Classroom.objects.create(name="旧版导入班", rows=1, cols=1)
        legacy_payload = {
            "meta": {"app": "不想排座位", "version": "1.0"},
            "classroom": {"name": "旧版导入班", "rows": 1, "cols": 1},
            "students": [
                {"name": "李四", "student_id": "1001", "gender": "M", "score": 90}
            ],
            "groups": [],
            "seats": [
                {
                    "row": 1,
                    "col": 1,
                    "cell_type": SeatCellType.SEAT,
                    "student_pk": None,
                    "student_id": "1001",
                    "student_name": "李四",
                    "group_name": None,
                }
            ],
            "constraints": [],
        }

        upload = SimpleUploadedFile(
            "legacy.seats",
            json.dumps(legacy_payload, ensure_ascii=False).encode("utf-8"),
            content_type="application/octet-stream",
        )
        response = self.client.post(
            reverse("import_seats_file", args=[classroom.pk]),
            {"seats_file": upload},
        )
        self.assertEqual(response.status_code, 302)

        student = Student.objects.get(classroom=classroom, student_id="1001")
        self.assertEqual((student.assigned_seat.row, student.assigned_seat.col), (1, 1))
        self.assertEqual(ClassroomHistoryEntry.objects.filter(classroom=classroom).count(), 1)

        undo_response = self.client.post(reverse("undo_action", args=[classroom.pk]))
        self.assertEqual(undo_response.status_code, 200)
        self.assertFalse(Student.objects.filter(classroom=classroom, student_id="1001").exists())


class GroupRotationTests(TestCase):
    def test_rotate_groups_swaps_group_positions_with_students(self):
        classroom = Classroom.objects.create(name="R1", rows=1, cols=4)
        g1 = SeatGroup.objects.create(classroom=classroom, name="G1", order=1)
        g2 = SeatGroup.objects.create(classroom=classroom, name="G2", order=2)

        seats = {col: classroom.seats.get(row=1, col=col) for col in [1, 2, 3, 4]}
        for col in [1, 2]:
            seats[col].group = g1
            seats[col].save(update_fields=["group"])
        for col in [3, 4]:
            seats[col].group = g2
            seats[col].save(update_fields=["group"])

        students = {}
        for idx, name in enumerate(["A", "B", "C", "D"], start=1):
            students[idx] = classroom.students.create(name=name, score=80 - idx)
            seats[idx].student = students[idx]
            seats[idx].save(update_fields=["student"])

        g1.leader = students[1]
        g1.save(update_fields=["leader"])

        url = reverse("rotate_groups", args=[classroom.pk])
        response = self.client.post(url, data=json.dumps({}), content_type="application/json")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json().get("status"), "success")

        seats = {col: classroom.seats.get(row=1, col=col) for col in [1, 2, 3, 4]}
        self.assertEqual(seats[1].group_id, g2.pk)
        self.assertEqual(seats[2].group_id, g2.pk)
        self.assertEqual(seats[3].group_id, g1.pk)
        self.assertEqual(seats[4].group_id, g1.pk)

        self.assertEqual(seats[1].student.name, "C")
        self.assertEqual(seats[2].student.name, "D")
        self.assertEqual(seats[3].student.name, "A")
        self.assertEqual(seats[4].student.name, "B")

        g1.refresh_from_db()
        self.assertEqual(g1.leader_id, students[1].pk)

    def test_rotate_groups_rejects_when_group_sizes_differ(self):
        classroom = Classroom.objects.create(name="R2", rows=1, cols=5)
        g1 = SeatGroup.objects.create(classroom=classroom, name="G1", order=1)
        g2 = SeatGroup.objects.create(classroom=classroom, name="G2", order=2)

        for col in [1, 2, 3]:
            seat = classroom.seats.get(row=1, col=col)
            seat.group = g1
            seat.save(update_fields=["group"])
        for col in [4, 5]:
            seat = classroom.seats.get(row=1, col=col)
            seat.group = g2
            seat.save(update_fields=["group"])

        url = reverse("rotate_groups", args=[classroom.pk])
        response = self.client.post(url, data=json.dumps({}), content_type="application/json")
        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.json().get("status"), "error")
        self.assertIn("座位数量不一致", response.json().get("message", ""))


class ConstraintManagementTests(TestCase):
    def setUp(self):
        self.classroom = Classroom.objects.create(name="约束测试班", rows=2, cols=3)
        self.student_a = self.classroom.students.create(name="张三")
        self.student_b = self.classroom.students.create(name="李四")

    def test_create_constraint_rejects_duplicate_enabled_rule(self):
        SeatConstraint.objects.create(
            classroom=self.classroom,
            constraint_type=SeatConstraint.ConstraintType.MUST_ROW,
            student=self.student_a,
            row=1,
            enabled=True,
        )

        response = self.client.post(
            reverse("create_constraint", args=[self.classroom.pk]),
            {
                "constraint_type": SeatConstraint.ConstraintType.MUST_ROW,
                "student_id": self.student_a.pk,
                "row": 1,
            },
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.json().get("status"), "error")
        self.assertIn("已存在相同约束", response.json().get("message", ""))

    def test_create_pair_constraint_normalizes_student_order(self):
        response = self.client.post(
            reverse("create_constraint", args=[self.classroom.pk]),
            {
                "constraint_type": SeatConstraint.ConstraintType.MUST_TOGETHER,
                "student_id": self.student_b.pk,
                "target_student_id": self.student_a.pk,
                "distance": 1,
            },
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 200)
        constraint = SeatConstraint.objects.get(classroom=self.classroom)
        self.assertEqual(constraint.student_id, self.student_a.pk)
        self.assertEqual(constraint.target_student_id, self.student_b.pk)

    def test_update_constraint_changes_rule_payload(self):
        constraint = SeatConstraint.objects.create(
            classroom=self.classroom,
            constraint_type=SeatConstraint.ConstraintType.MUST_ROW,
            student=self.student_a,
            row=1,
            enabled=True,
        )

        response = self.client.post(
            reverse("update_constraint", args=[self.classroom.pk, constraint.pk]),
            {
                "constraint_type": SeatConstraint.ConstraintType.MUST_COL,
                "student_id": self.student_a.pk,
                "col": 2,
                "note": "改成列约束",
                "enabled": "1",
            },
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 200)
        constraint.refresh_from_db()
        self.assertEqual(constraint.constraint_type, SeatConstraint.ConstraintType.MUST_COL)
        self.assertEqual(constraint.col, 2)
        self.assertIsNone(constraint.row)
        self.assertEqual(constraint.note, "改成列约束")

    def test_toggle_constraint_rejects_enable_when_conflicting(self):
        SeatConstraint.objects.create(
            classroom=self.classroom,
            constraint_type=SeatConstraint.ConstraintType.MUST_ROW,
            student=self.student_a,
            row=1,
            enabled=True,
        )
        constraint = SeatConstraint.objects.create(
            classroom=self.classroom,
            constraint_type=SeatConstraint.ConstraintType.MUST_ROW,
            student=self.student_a,
            row=2,
            enabled=False,
        )

        response = self.client.post(
            reverse("toggle_constraint", args=[self.classroom.pk, constraint.pk]),
            {"enabled": "1"},
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.json().get("status"), "error")
        self.assertIn("不能同时指定多个不同行", response.json().get("message", ""))

    def test_classroom_state_returns_constraints_and_issue_metrics(self):
        SeatConstraint.objects.create(
            classroom=self.classroom,
            constraint_type=SeatConstraint.ConstraintType.MUST_SEAT,
            student=self.student_a,
            row=1,
            col=1,
            enabled=True,
        )
        SeatConstraint.objects.create(
            classroom=self.classroom,
            constraint_type=SeatConstraint.ConstraintType.FORBID_SEAT,
            student=self.student_a,
            row=1,
            col=1,
            enabled=True,
        )

        response = self.client.get(
            reverse("classroom_state", args=[self.classroom.pk]),
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(len(payload.get("constraints", [])), 2)
        self.assertEqual((payload.get("constraint_metrics") or {}).get("with_issues"), 2)
        self.assertTrue(all(item.get("issue_count", 0) > 0 for item in payload.get("constraints", [])))


class LayoutShiftTests(TestCase):
    def test_shift_layout_right_keeps_both_podium_guards_in_place(self):
        classroom = Classroom.objects.create(name="LS-Guard-LR", rows=1, cols=4)
        left_guardian = classroom.students.create(name="左护法")
        middle_a = classroom.students.create(name="中间甲")
        middle_b = classroom.students.create(name="中间乙")
        right_guardian = classroom.students.create(name="右护法")

        classroom.seats.filter(row=1, col=1).update(student=left_guardian)
        classroom.seats.filter(row=1, col=2).update(student=middle_a)
        classroom.seats.filter(row=1, col=3).update(student=middle_b)
        classroom.seats.filter(row=1, col=4).update(student=right_guardian)
        classroom.left_guardian = left_guardian
        classroom.right_guardian = right_guardian
        classroom.save(update_fields=["left_guardian", "right_guardian"])

        response = self.client.post(
            reverse("shift_layout", args=[classroom.pk]),
            data=json.dumps({"direction": "right", "steps": 1}),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json().get("status"), "success")

        left_guardian.refresh_from_db()
        middle_a.refresh_from_db()
        middle_b.refresh_from_db()
        right_guardian.refresh_from_db()
        self.assertEqual((left_guardian.assigned_seat.row, left_guardian.assigned_seat.col), (1, 1))
        self.assertEqual((right_guardian.assigned_seat.row, right_guardian.assigned_seat.col), (1, 4))
        self.assertEqual((middle_a.assigned_seat.row, middle_a.assigned_seat.col), (1, 3))
        self.assertEqual((middle_b.assigned_seat.row, middle_b.assigned_seat.col), (1, 2))

    def test_shift_layout_right_wraps_layout_and_preserves_layout_data(self):
        classroom = Classroom.objects.create(name="LS1", rows=2, cols=3)
        student = classroom.students.create(name="张三")
        group = SeatGroup.objects.create(classroom=classroom, name="第1组", order=1)

        seat_a = classroom.seats.get(row=1, col=1)
        seat_a.student = student
        seat_a.group = group
        seat_a.save(update_fields=["student", "group"])

        seat_b = classroom.seats.get(row=1, col=2)
        seat_b.cell_type = SeatCellType.AISLE
        seat_b.save(update_fields=["cell_type"])

        SeatConstraint.objects.create(
            classroom=classroom,
            constraint_type=SeatConstraint.ConstraintType.MUST_SEAT,
            student=student,
            row=1,
            col=1,
        )

        url = reverse("shift_layout", args=[classroom.pk])
        response = self.client.post(
            url,
            data=json.dumps({"direction": "right", "steps": 2}),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json().get("status"), "success")

        classroom.refresh_from_db()
        self.assertEqual(classroom.cols, 3)

        shifted = classroom.seats.get(row=1, col=3)
        self.assertEqual(shifted.student_id, student.pk)
        self.assertEqual(shifted.group_id, group.pk)
        self.assertEqual(classroom.seats.get(row=1, col=1).cell_type, SeatCellType.AISLE)

        constraint = classroom.constraints.get(student=student)
        self.assertEqual((constraint.row, constraint.col), (1, 3))

    def test_shift_layout_right_keeps_podium_guardian_constraint_in_place(self):
        classroom = Classroom.objects.create(name="LS-Guard-Constraint", rows=1, cols=3)
        guardian = classroom.students.create(name="护法同学")
        other_student = classroom.students.create(name="普通同学")

        classroom.seats.filter(row=1, col=1).update(student=guardian)
        classroom.seats.filter(row=1, col=2).update(student=other_student)
        classroom.left_guardian = guardian
        classroom.save(update_fields=["left_guardian"])

        SeatConstraint.objects.create(
            classroom=classroom,
            constraint_type=SeatConstraint.ConstraintType.MUST_SEAT,
            student=guardian,
            row=1,
            col=1,
        )

        response = self.client.post(
            reverse("shift_layout", args=[classroom.pk]),
            data=json.dumps({"direction": "right", "steps": 1}),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        guardian.refresh_from_db()
        self.assertEqual((guardian.assigned_seat.row, guardian.assigned_seat.col), (1, 1))

        constraint = classroom.constraints.get(student=guardian)
        self.assertEqual((constraint.row, constraint.col), (1, 1))

    def test_shift_layout_right_keeps_manually_fixed_student_in_place(self):
        classroom = Classroom.objects.create(name="LS-Fixed-Seat", rows=1, cols=4)
        fixed_student = classroom.students.create(name="固定同学")
        other_student = classroom.students.create(name="轮换同学")

        classroom.seats.filter(row=1, col=2).update(student=fixed_student)
        classroom.seats.filter(row=1, col=3).update(student=other_student)

        fixed_response = self.client.post(
            reverse("toggle_fixed_seat", args=[classroom.pk]),
            data=json.dumps({"row": 1, "col": 2, "enabled": True}),
            content_type="application/json",
        )
        self.assertEqual(fixed_response.status_code, 200)

        response = self.client.post(
            reverse("shift_layout", args=[classroom.pk]),
            data=json.dumps({"direction": "right", "steps": 1}),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)

        fixed_student.refresh_from_db()
        other_student.refresh_from_db()
        self.assertEqual((fixed_student.assigned_seat.row, fixed_student.assigned_seat.col), (1, 2))
        self.assertEqual((other_student.assigned_seat.row, other_student.assigned_seat.col), (1, 4))

    def test_shift_layout_left_wraps_and_supports_undo_and_redo(self):
        classroom = Classroom.objects.create(name="LS2", rows=1, cols=4)
        student = classroom.students.create(name="李四")

        seat = classroom.seats.get(row=1, col=1)
        seat.student = student
        seat.save(update_fields=["student"])

        classroom.seats.filter(row=1, col=3).update(cell_type=SeatCellType.PODIUM)

        shift_url = reverse("shift_layout", args=[classroom.pk])
        response = self.client.post(
            shift_url,
            data=json.dumps({"direction": "left", "steps": 1}),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        classroom.refresh_from_db()
        student.refresh_from_db()
        self.assertEqual(classroom.cols, 4)
        self.assertEqual((student.assigned_seat.row, student.assigned_seat.col), (1, 4))
        self.assertEqual(classroom.seats.get(row=1, col=3).cell_type, SeatCellType.PODIUM)

        undo_url = reverse("undo_action", args=[classroom.pk])
        redo_url = reverse("redo_action", args=[classroom.pk])

        undo_response = self.client.post(undo_url)
        self.assertEqual(undo_response.status_code, 200)
        classroom.refresh_from_db()
        student.refresh_from_db()
        self.assertEqual(classroom.cols, 4)
        self.assertEqual((student.assigned_seat.row, student.assigned_seat.col), (1, 1))
        self.assertEqual(classroom.seats.get(row=1, col=3).cell_type, SeatCellType.PODIUM)

        redo_response = self.client.post(redo_url)
        self.assertEqual(redo_response.status_code, 200)
        classroom.refresh_from_db()
        student.refresh_from_db()
        self.assertEqual(classroom.cols, 4)
        self.assertEqual((student.assigned_seat.row, student.assigned_seat.col), (1, 4))

    def test_shift_layout_left_wraps_non_empty_leading_columns(self):
        classroom = Classroom.objects.create(name="LS3", rows=1, cols=3)
        student = classroom.students.create(name="王五")
        seat = classroom.seats.get(row=1, col=1)
        seat.student = student
        seat.save(update_fields=["student"])

        url = reverse("shift_layout", args=[classroom.pk])

        response = self.client.post(
            url,
            data=json.dumps({"direction": "left", "steps": 1}),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        classroom.refresh_from_db()
        student.refresh_from_db()
        self.assertEqual((student.assigned_seat.row, student.assigned_seat.col), (1, 3))

    def test_shift_layout_left_wraps_column_constraints(self):
        classroom = Classroom.objects.create(name="LS4", rows=1, cols=3)
        student = classroom.students.create(name="王五")

        SeatConstraint.objects.create(
            classroom=classroom,
            constraint_type=SeatConstraint.ConstraintType.MUST_COL,
            student=student,
            col=1,
        )

        url = reverse("shift_layout", args=[classroom.pk])
        response = self.client.post(
            url,
            data=json.dumps({"direction": "left", "steps": 1}),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        constraint = classroom.constraints.get(student=student)
        self.assertEqual(constraint.col, 3)

    def test_shift_layout_back_wraps_layout_and_preserves_layout_data(self):
        classroom = Classroom.objects.create(name="LS5", rows=3, cols=2)
        student = classroom.students.create(name="赵六")
        group = SeatGroup.objects.create(classroom=classroom, name="第2组", order=2)

        seat_a = classroom.seats.get(row=1, col=1)
        seat_a.student = student
        seat_a.group = group
        seat_a.save(update_fields=["student", "group"])

        seat_b = classroom.seats.get(row=2, col=1)
        seat_b.cell_type = SeatCellType.AISLE
        seat_b.save(update_fields=["cell_type"])

        SeatConstraint.objects.create(
            classroom=classroom,
            constraint_type=SeatConstraint.ConstraintType.MUST_SEAT,
            student=student,
            row=1,
            col=1,
        )

        url = reverse("shift_layout", args=[classroom.pk])
        response = self.client.post(
            url,
            data=json.dumps({"direction": "back", "steps": 2}),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json().get("status"), "success")

        classroom.refresh_from_db()
        self.assertEqual(classroom.rows, 3)

        shifted = classroom.seats.get(row=3, col=1)
        self.assertEqual(shifted.student_id, student.pk)
        self.assertEqual(shifted.group_id, group.pk)
        self.assertEqual(classroom.seats.get(row=1, col=1).cell_type, SeatCellType.AISLE)

        constraint = classroom.constraints.get(student=student)
        self.assertEqual((constraint.row, constraint.col), (3, 1))

    def test_shift_layout_front_wraps_and_supports_undo_and_redo(self):
        classroom = Classroom.objects.create(name="LS6", rows=4, cols=1)
        student = classroom.students.create(name="孙七")

        seat = classroom.seats.get(row=2, col=1)
        seat.student = student
        seat.save(update_fields=["student"])

        classroom.seats.filter(row=3, col=1).update(cell_type=SeatCellType.PODIUM)

        shift_url = reverse("shift_layout", args=[classroom.pk])
        response = self.client.post(
            shift_url,
            data=json.dumps({"direction": "front", "steps": 1}),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        classroom.refresh_from_db()
        student.refresh_from_db()
        self.assertEqual(classroom.rows, 4)
        self.assertEqual((student.assigned_seat.row, student.assigned_seat.col), (1, 1))
        self.assertEqual(classroom.seats.get(row=3, col=1).cell_type, SeatCellType.PODIUM)

        undo_url = reverse("undo_action", args=[classroom.pk])
        redo_url = reverse("redo_action", args=[classroom.pk])

        undo_response = self.client.post(undo_url)
        self.assertEqual(undo_response.status_code, 200)
        classroom.refresh_from_db()
        student.refresh_from_db()
        self.assertEqual(classroom.rows, 4)
        self.assertEqual((student.assigned_seat.row, student.assigned_seat.col), (2, 1))
        self.assertEqual(classroom.seats.get(row=3, col=1).cell_type, SeatCellType.PODIUM)

        redo_response = self.client.post(redo_url)
        self.assertEqual(redo_response.status_code, 200)
        classroom.refresh_from_db()
        student.refresh_from_db()
        self.assertEqual(classroom.rows, 4)
        self.assertEqual((student.assigned_seat.row, student.assigned_seat.col), (1, 1))

    def test_shift_layout_front_keeps_podium_guardian_in_place(self):
        classroom = Classroom.objects.create(name="LS-Guard-FB", rows=3, cols=1)
        guardian = classroom.students.create(name="前排护法")
        middle_student = classroom.students.create(name="中排同学")
        back_student = classroom.students.create(name="后排同学")

        classroom.seats.filter(row=1, col=1).update(student=guardian)
        classroom.seats.filter(row=2, col=1).update(student=middle_student)
        classroom.seats.filter(row=3, col=1).update(student=back_student)
        classroom.left_guardian = guardian
        classroom.save(update_fields=["left_guardian"])

        response = self.client.post(
            reverse("shift_layout", args=[classroom.pk]),
            data=json.dumps({"direction": "front", "steps": 1}),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json().get("status"), "success")

        guardian.refresh_from_db()
        middle_student.refresh_from_db()
        back_student.refresh_from_db()
        self.assertEqual((guardian.assigned_seat.row, guardian.assigned_seat.col), (1, 1))
        self.assertEqual((middle_student.assigned_seat.row, middle_student.assigned_seat.col), (3, 1))
        self.assertEqual((back_student.assigned_seat.row, back_student.assigned_seat.col), (2, 1))

    def test_shift_layout_front_skips_rows_containing_podium(self):
        classroom = Classroom.objects.create(name="LS-Front-Skip-Podium-Row", rows=4, cols=2)
        front_student = classroom.students.create(name="前排同学")
        middle_student = classroom.students.create(name="中排同学")
        podium_row_student = classroom.students.create(name="讲台同行")
        back_student = classroom.students.create(name="后排同学")

        classroom.seats.filter(row=1, col=1).update(student=front_student)
        classroom.seats.filter(row=2, col=1).update(student=middle_student)
        classroom.seats.filter(row=3, col=1).update(cell_type=SeatCellType.PODIUM, student=None, group=None)
        classroom.seats.filter(row=3, col=2).update(student=podium_row_student)
        classroom.seats.filter(row=4, col=1).update(student=back_student)

        SeatConstraint.objects.create(
            classroom=classroom,
            constraint_type=SeatConstraint.ConstraintType.MUST_ROW,
            student=middle_student,
            row=2,
        )
        SeatConstraint.objects.create(
            classroom=classroom,
            constraint_type=SeatConstraint.ConstraintType.MUST_ROW,
            student=podium_row_student,
            row=3,
        )

        response = self.client.post(
            reverse("shift_layout", args=[classroom.pk]),
            data=json.dumps({"direction": "front", "steps": 1}),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)

        front_student.refresh_from_db()
        middle_student.refresh_from_db()
        podium_row_student.refresh_from_db()
        back_student.refresh_from_db()
        self.assertEqual((front_student.assigned_seat.row, front_student.assigned_seat.col), (4, 1))
        self.assertEqual((middle_student.assigned_seat.row, middle_student.assigned_seat.col), (1, 1))
        self.assertEqual((podium_row_student.assigned_seat.row, podium_row_student.assigned_seat.col), (3, 2))
        self.assertEqual((back_student.assigned_seat.row, back_student.assigned_seat.col), (2, 1))
        self.assertEqual(classroom.seats.get(row=3, col=1).cell_type, SeatCellType.PODIUM)

        middle_constraint = classroom.constraints.get(student=middle_student)
        podium_row_constraint = classroom.constraints.get(student=podium_row_student)
        self.assertEqual(middle_constraint.row, 1)
        self.assertEqual(podium_row_constraint.row, 3)

    def test_shift_layout_back_skips_rows_containing_podium(self):
        classroom = Classroom.objects.create(name="LS-Back-Skip-Podium-Row", rows=4, cols=2)
        front_student = classroom.students.create(name="前排同学")
        middle_student = classroom.students.create(name="中排同学")
        podium_row_student = classroom.students.create(name="讲台同行")
        back_student = classroom.students.create(name="后排同学")

        classroom.seats.filter(row=1, col=1).update(student=front_student)
        classroom.seats.filter(row=2, col=1).update(student=middle_student)
        classroom.seats.filter(row=3, col=1).update(cell_type=SeatCellType.PODIUM, student=None, group=None)
        classroom.seats.filter(row=3, col=2).update(student=podium_row_student)
        classroom.seats.filter(row=4, col=1).update(student=back_student)

        response = self.client.post(
            reverse("shift_layout", args=[classroom.pk]),
            data=json.dumps({"direction": "back", "steps": 1}),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)

        front_student.refresh_from_db()
        middle_student.refresh_from_db()
        podium_row_student.refresh_from_db()
        back_student.refresh_from_db()
        self.assertEqual((front_student.assigned_seat.row, front_student.assigned_seat.col), (2, 1))
        self.assertEqual((middle_student.assigned_seat.row, middle_student.assigned_seat.col), (4, 1))
        self.assertEqual((podium_row_student.assigned_seat.row, podium_row_student.assigned_seat.col), (3, 2))
        self.assertEqual((back_student.assigned_seat.row, back_student.assigned_seat.col), (1, 1))
        self.assertEqual(classroom.seats.get(row=3, col=1).cell_type, SeatCellType.PODIUM)

    def test_shift_layout_front_wraps_non_empty_leading_rows(self):
        classroom = Classroom.objects.create(name="LS7", rows=3, cols=1)
        student = classroom.students.create(name="周八")
        seat = classroom.seats.get(row=1, col=1)
        seat.student = student
        seat.save(update_fields=["student"])

        url = reverse("shift_layout", args=[classroom.pk])

        response = self.client.post(
            url,
            data=json.dumps({"direction": "front", "steps": 1}),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        classroom.refresh_from_db()
        student.refresh_from_db()
        self.assertEqual((student.assigned_seat.row, student.assigned_seat.col), (3, 1))

    def test_shift_layout_front_wraps_row_constraints(self):
        classroom = Classroom.objects.create(name="LS8", rows=3, cols=1)
        student = classroom.students.create(name="周八")

        SeatConstraint.objects.create(
            classroom=classroom,
            constraint_type=SeatConstraint.ConstraintType.MUST_ROW,
            student=student,
            row=1,
        )

        url = reverse("shift_layout", args=[classroom.pk])
        response = self.client.post(
            url,
            data=json.dumps({"direction": "front", "steps": 1}),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        constraint = classroom.constraints.get(student=student)
        self.assertEqual(constraint.row, 3)

    def test_shift_layout_left_uses_template_blocks_for_2_1_2(self):
        classroom = Classroom.objects.create(name="LS9", rows=2, cols=5)
        left_student = classroom.students.create(name="左组学生")
        right_student = classroom.students.create(name="右组学生")
        left_group = SeatGroup.objects.create(classroom=classroom, name="左组", order=1)
        right_group = SeatGroup.objects.create(classroom=classroom, name="右组", order=2)

        left_seat = classroom.seats.get(row=1, col=1)
        left_seat.student = left_student
        left_seat.group = left_group
        left_seat.save(update_fields=["student", "group"])

        right_seat = classroom.seats.get(row=1, col=4)
        right_seat.student = right_student
        right_seat.group = right_group
        right_seat.save(update_fields=["student", "group"])

        classroom.seats.filter(col=3).update(cell_type=SeatCellType.AISLE, student=None, group=None)
        classroom.seats.filter(row=2, col=2).update(cell_type=SeatCellType.AISLE, student=None, group=None)

        SeatConstraint.objects.create(
            classroom=classroom,
            constraint_type=SeatConstraint.ConstraintType.MUST_SEAT,
            student=left_student,
            row=1,
            col=1,
        )

        url = reverse("shift_layout", args=[classroom.pk])
        response = self.client.post(
            url,
            data=json.dumps({"direction": "left", "steps": 1}),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload.get("shift_mode"), "template")
        self.assertEqual(payload.get("template_signature"), "2+1+2")
        self.assertIn("布局模板", payload.get("message", ""))

        left_student.refresh_from_db()
        right_student.refresh_from_db()

        self.assertEqual(left_student.assigned_seat.col, 4)
        self.assertEqual(right_student.assigned_seat.col, 1)
        self.assertEqual(classroom.seats.get(row=1, col=3).cell_type, SeatCellType.AISLE)
        self.assertEqual(classroom.seats.get(row=2, col=2).cell_type, SeatCellType.SEAT)
        self.assertEqual(classroom.seats.get(row=2, col=5).cell_type, SeatCellType.AISLE)

        constraint = classroom.constraints.get(student=left_student)
        self.assertEqual((constraint.row, constraint.col), (1, 4))

    def test_shift_layout_right_uses_template_blocks_for_2_1_2_1_2(self):
        classroom = Classroom.objects.create(name="LS10", rows=1, cols=8)
        student_a = classroom.students.create(name="A")
        student_b = classroom.students.create(name="B")
        student_c = classroom.students.create(name="C")

        classroom.seats.filter(row=1, col=1).update(student=student_a)
        classroom.seats.filter(row=1, col=4).update(student=student_b)
        classroom.seats.filter(row=1, col=7).update(student=student_c)
        classroom.seats.filter(col__in=[3, 6]).update(cell_type=SeatCellType.AISLE, student=None, group=None)

        url = reverse("shift_layout", args=[classroom.pk])
        response = self.client.post(
            url,
            data=json.dumps({"direction": "right", "steps": 1}),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload.get("shift_mode"), "template")
        self.assertEqual(payload.get("template_signature"), "2+1+2+1+2")

        student_a.refresh_from_db()
        student_b.refresh_from_db()
        student_c.refresh_from_db()

        self.assertEqual(student_c.assigned_seat.col, 1)
        self.assertEqual(student_a.assigned_seat.col, 4)
        self.assertEqual(student_b.assigned_seat.col, 7)
        self.assertEqual(classroom.seats.get(row=1, col=3).cell_type, SeatCellType.AISLE)
        self.assertEqual(classroom.seats.get(row=1, col=6).cell_type, SeatCellType.AISLE)

    def test_shift_layout_right_supports_different_seat_block_widths(self):
        classroom = Classroom.objects.create(name="LS11", rows=1, cols=8)
        left_student = classroom.students.create(name="左侧")
        right_student = classroom.students.create(name="右侧")

        classroom.seats.filter(row=1, col=3).update(student=left_student)
        classroom.seats.filter(row=1, col=5).update(student=right_student)
        classroom.seats.filter(col=4).update(cell_type=SeatCellType.AISLE, student=None, group=None)

        url = reverse("shift_layout", args=[classroom.pk])
        response = self.client.post(
            url,
            data=json.dumps({"direction": "right", "steps": 1}),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload.get("shift_mode"), "template")
        self.assertEqual(payload.get("template_signature"), "3+1+4")

        left_student.refresh_from_db()
        right_student.refresh_from_db()

        self.assertEqual(right_student.assigned_seat.col, 1)
        self.assertEqual(left_student.assigned_seat.col, 8)
        self.assertEqual(classroom.seats.get(row=1, col=5).cell_type, SeatCellType.AISLE)
        self.assertEqual(classroom.seats.get(row=1, col=4).cell_type, SeatCellType.SEAT)

    def test_shift_layout_left_rotates_single_columns_when_large_groups_disabled(self):
        classroom = Classroom.objects.create(name="LS11-C1", rows=1, cols=7)
        student_a1 = classroom.students.create(name="a1")
        student_a2 = classroom.students.create(name="a2")
        student_a3 = classroom.students.create(name="a3")
        student_a4 = classroom.students.create(name="a4")

        classroom.seats.filter(row=1, col=2).update(student=student_a1)
        classroom.seats.filter(row=1, col=4).update(student=student_a2)
        classroom.seats.filter(row=1, col=5).update(student=student_a3)
        classroom.seats.filter(row=1, col=7).update(student=student_a4)
        classroom.seats.filter(col__in=[1, 3, 6]).update(cell_type=SeatCellType.AISLE, student=None, group=None)

        SeatConstraint.objects.create(
            classroom=classroom,
            constraint_type=SeatConstraint.ConstraintType.MUST_SEAT,
            student=student_a2,
            row=1,
            col=4,
        )

        url = reverse("shift_layout", args=[classroom.pk])
        response = self.client.post(
            url,
            data=json.dumps({"direction": "left", "steps": 1, "use_large_groups": False}),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload.get("shift_mode"), "column")
        self.assertEqual(payload.get("seat_column_count"), 4)

        student_a1.refresh_from_db()
        student_a2.refresh_from_db()
        student_a3.refresh_from_db()
        student_a4.refresh_from_db()

        self.assertEqual(student_a4.assigned_seat.col, 2)
        self.assertEqual(student_a1.assigned_seat.col, 4)
        self.assertEqual(student_a2.assigned_seat.col, 5)
        self.assertEqual(student_a3.assigned_seat.col, 7)
        self.assertEqual(classroom.seats.get(row=1, col=1).cell_type, SeatCellType.AISLE)
        self.assertEqual(classroom.seats.get(row=1, col=3).cell_type, SeatCellType.AISLE)
        self.assertEqual(classroom.seats.get(row=1, col=6).cell_type, SeatCellType.AISLE)

        constraint = classroom.constraints.get(student=student_a2)
        self.assertEqual((constraint.row, constraint.col), (1, 5))

    def test_shift_layout_right_rotates_single_columns_when_large_groups_disabled(self):
        classroom = Classroom.objects.create(name="LS11-C2", rows=1, cols=7)
        student_a1 = classroom.students.create(name="a1")
        student_a2 = classroom.students.create(name="a2")
        student_a3 = classroom.students.create(name="a3")
        student_a4 = classroom.students.create(name="a4")

        classroom.seats.filter(row=1, col=2).update(student=student_a1)
        classroom.seats.filter(row=1, col=4).update(student=student_a2)
        classroom.seats.filter(row=1, col=5).update(student=student_a3)
        classroom.seats.filter(row=1, col=7).update(student=student_a4)
        classroom.seats.filter(col__in=[1, 3, 6]).update(cell_type=SeatCellType.AISLE, student=None, group=None)

        url = reverse("shift_layout", args=[classroom.pk])
        response = self.client.post(
            url,
            data=json.dumps({"direction": "right", "steps": 1, "use_large_groups": False}),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload.get("shift_mode"), "column")
        self.assertEqual(payload.get("seat_column_count"), 4)

        student_a1.refresh_from_db()
        student_a2.refresh_from_db()
        student_a3.refresh_from_db()
        student_a4.refresh_from_db()

        self.assertEqual(student_a2.assigned_seat.col, 2)
        self.assertEqual(student_a3.assigned_seat.col, 4)
        self.assertEqual(student_a4.assigned_seat.col, 5)
        self.assertEqual(student_a1.assigned_seat.col, 7)

    def test_shift_layout_left_preserves_symmetric_structure_for_1_1_3_1_3_1_1(self):
        classroom = Classroom.objects.create(name="LS11-S", rows=1, cols=11)
        left_edge_student = classroom.students.create(name="左侧单列")
        left_middle_student = classroom.students.create(name="左侧三列")
        right_middle_student = classroom.students.create(name="右侧三列")
        right_edge_student = classroom.students.create(name="右侧单列")

        classroom.seats.filter(row=1, col=1).update(student=left_edge_student)
        classroom.seats.filter(row=1, col=3).update(student=left_middle_student)
        classroom.seats.filter(row=1, col=7).update(student=right_middle_student)
        classroom.seats.filter(row=1, col=11).update(student=right_edge_student)
        classroom.seats.filter(col__in=[2, 6, 10]).update(cell_type=SeatCellType.AISLE, student=None, group=None)

        SeatConstraint.objects.create(
            classroom=classroom,
            constraint_type=SeatConstraint.ConstraintType.MUST_SEAT,
            student=left_middle_student,
            row=1,
            col=3,
        )

        url = reverse("shift_layout", args=[classroom.pk])
        response = self.client.post(
            url,
            data=json.dumps({"direction": "left", "steps": 1}),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload.get("shift_mode"), "template")
        self.assertEqual(payload.get("template_signature"), "1+1+3+1+3+1+1")

        left_edge_student.refresh_from_db()
        left_middle_student.refresh_from_db()
        right_middle_student.refresh_from_db()
        right_edge_student.refresh_from_db()

        self.assertEqual(right_edge_student.assigned_seat.col, 1)
        self.assertEqual(right_middle_student.assigned_seat.col, 3)
        self.assertEqual(left_middle_student.assigned_seat.col, 7)
        self.assertEqual(left_edge_student.assigned_seat.col, 11)
        self.assertEqual(classroom.seats.get(row=1, col=2).cell_type, SeatCellType.AISLE)
        self.assertEqual(classroom.seats.get(row=1, col=6).cell_type, SeatCellType.AISLE)
        self.assertEqual(classroom.seats.get(row=1, col=10).cell_type, SeatCellType.AISLE)
        self.assertEqual(classroom.seats.get(row=1, col=1).cell_type, SeatCellType.SEAT)
        self.assertEqual(classroom.seats.get(row=1, col=11).cell_type, SeatCellType.SEAT)

        constraint = classroom.constraints.get(student=left_middle_student)
        self.assertEqual((constraint.row, constraint.col), (1, 7))

    def test_shift_layout_left_falls_back_when_template_is_not_recognized(self):
        classroom = Classroom.objects.create(name="LS12", rows=1, cols=3)
        student = classroom.students.create(name="普通移动")
        classroom.seats.filter(row=1, col=1).update(student=student)

        url = reverse("shift_layout", args=[classroom.pk])
        response = self.client.post(
            url,
            data=json.dumps({"direction": "left", "steps": 1}),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload.get("shift_mode"), "normal")
        self.assertIn("回退为普通左移", payload.get("message", ""))
        self.assertTrue(payload.get("fallback_reason"))

        student.refresh_from_db()
        self.assertEqual(student.assigned_seat.col, 3)


class StudentImportTests(TestCase):
    @staticmethod
    def _build_excel_file(filename="students.xlsx"):
        workbook = openpyxl.Workbook()
        workbook.active.title = "说明"
        workbook.active.append(["导入说明"])
        score_sheet = workbook.create_sheet("成绩表")
        score_sheet.append(["2026 学年"])
        score_sheet.append([])
        score_sheet.append(["学生编号", "学生姓名", "男女", "期末成绩"])
        score_sheet.append(["S001", "张三", "男", 95])
        score_sheet.append(["S002", "李四", "女", 88])
        output = BytesIO()
        workbook.save(output)
        workbook.close()
        return SimpleUploadedFile(
            filename,
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def test_excel_upload_always_waits_for_visual_confirmation(self):
        classroom = Classroom.objects.create(name="可视化导入", rows=2, cols=2)
        classroom.students.create(name="原学生", score=60)
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.append(["姓名", "总分"])
        worksheet.append(["新学生", 99])
        output = BytesIO()
        workbook.save(output)
        workbook.close()
        upload = SimpleUploadedFile(
            "standard.xlsx",
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        with tempfile.TemporaryDirectory() as temp_dir, override_settings(BASE_DIR=temp_dir):
            response = self.client.post(
                reverse("import_students", args=[classroom.pk]),
                {"action": "upload", "excel_file": upload, "import_mode": "replace"},
            )

            self.assertEqual(response.status_code, 200)
            payload = response.json()
            self.assertEqual(payload["status"], "ready")
            self.assertEqual(payload["preview_data"][0], ["姓名", "总分"])
            self.assertEqual(payload["suggested"]["start_row"], 2)
            self.assertEqual(payload["suggested"]["name_col_index"], 0)
            self.assertEqual(payload["suggested"]["score_col_index"], 1)
            self.assertEqual(classroom.students.count(), 1)
            self.assertTrue(classroom.students.filter(name="原学生").exists())
            self.assertTrue((Path(temp_dir) / "temp_imports" / f'{payload["file_id"]}.xlsx').exists())

    def test_excel_import_supports_sheet_row_and_all_column_mappings(self):
        classroom = Classroom.objects.create(name="多工作表导入", rows=2, cols=2)
        upload = self._build_excel_file()

        with tempfile.TemporaryDirectory() as temp_dir, override_settings(BASE_DIR=temp_dir):
            url = reverse("import_students", args=[classroom.pk])
            upload_response = self.client.post(url, {"action": "upload", "excel_file": upload})
            self.assertEqual(upload_response.status_code, 200)
            file_id = upload_response.json()["file_id"]

            preview_response = self.client.post(
                url,
                {"action": "preview", "file_id": file_id, "sheet_name": "成绩表"},
            )
            self.assertEqual(preview_response.status_code, 200)
            preview = preview_response.json()
            self.assertEqual(preview["sheet_name"], "成绩表")
            self.assertEqual(preview["total_rows"], 5)
            self.assertEqual(preview["total_cols"], 4)

            confirm_response = self.client.post(
                url,
                {
                    "action": "confirm",
                    "file_id": file_id,
                    "sheet_name": "成绩表",
                    "start_row": "4",
                    "name_col_index": "1",
                    "student_id_col_index": "0",
                    "gender_col_index": "2",
                    "score_col_index": "3",
                    "import_mode": "replace",
                },
            )

            self.assertEqual(confirm_response.status_code, 200)
            self.assertEqual(confirm_response.json()["status"], "success")
            self.assertEqual(classroom.students.count(), 2)
            zhangsan = classroom.students.get(student_id="S001")
            lisi = classroom.students.get(student_id="S002")
            self.assertEqual((zhangsan.name, zhangsan.gender, zhangsan.score), ("张三", "M", 95))
            self.assertEqual((lisi.name, lisi.gender, lisi.score), ("李四", "F", 88))
            self.assertFalse((Path(temp_dir) / "temp_imports" / f"{file_id}.xlsx").exists())

    def test_excel_import_rejects_duplicate_field_columns(self):
        classroom = Classroom.objects.create(name="重复列导入", rows=1, cols=1)
        upload = self._build_excel_file()

        with tempfile.TemporaryDirectory() as temp_dir, override_settings(BASE_DIR=temp_dir):
            url = reverse("import_students", args=[classroom.pk])
            file_id = self.client.post(url, {"action": "upload", "excel_file": upload}).json()["file_id"]
            response = self.client.post(
                url,
                {
                    "action": "confirm",
                    "file_id": file_id,
                    "sheet_name": "成绩表",
                    "start_row": "4",
                    "name_col_index": "1",
                    "score_col_index": "1",
                },
            )

            self.assertEqual(response.status_code, 400)
            self.assertIn("同一列不能绑定多个字段", response.json()["message"])
            self.assertEqual(classroom.students.count(), 0)

    def test_excel_import_discard_removes_temporary_file(self):
        classroom = Classroom.objects.create(name="取消导入", rows=1, cols=1)
        upload = self._build_excel_file()

        with tempfile.TemporaryDirectory() as temp_dir, override_settings(BASE_DIR=temp_dir):
            url = reverse("import_students", args=[classroom.pk])
            file_id = self.client.post(url, {"action": "upload", "excel_file": upload}).json()["file_id"]
            temp_file = Path(temp_dir) / "temp_imports" / f"{file_id}.xlsx"
            self.assertTrue(temp_file.exists())

            response = self.client.post(url, {"action": "discard", "file_id": file_id})

            self.assertEqual(response.status_code, 200)
            self.assertEqual(response.json()["status"], "success")
            self.assertFalse(temp_file.exists())

    def test_process_import_match_updates_existing_students(self):
        classroom = Classroom.objects.create(name="导入匹配", rows=2, cols=2)
        stu_by_id = classroom.students.create(name="张三", student_id="1001", score=60)
        stu_by_name = classroom.students.create(name="李四", score=55)

        df = pd.DataFrame(
            [
                {"姓名": "张三", "学号": "1001", "总分": 95},
                {"姓名": "李四", "总分": 88},
                {"姓名": "王五", "总分": 77},
            ]
        )

        result = _process_import(
            classroom,
            df,
            "姓名",
            "学号",
            None,
            "总分",
            import_mode=IMPORT_MODE_MATCH,
        )

        stu_by_id.refresh_from_db()
        stu_by_name.refresh_from_db()
        self.assertEqual(stu_by_id.score, 95)
        self.assertEqual(stu_by_name.score, 88)
        self.assertEqual(classroom.students.count(), 3)
        self.assertTrue(classroom.students.filter(name="王五", score=77).exists())
        self.assertEqual(result["updated"], 2)
        self.assertEqual(result["created"], 1)
        self.assertEqual(result["skipped"], 0)

    def test_process_import_replace_rebuilds_students(self):
        classroom = Classroom.objects.create(name="导入清空", rows=2, cols=2)
        classroom.students.create(name="旧学生", student_id="A01", score=30)

        df = pd.DataFrame(
            [
                {"姓名": "新学生1", "学号": "N01", "总分": 91},
                {"姓名": "新学生2", "学号": "N02", "总分": 85},
            ]
        )

        result = _process_import(
            classroom,
            df,
            "姓名",
            "学号",
            None,
            "总分",
            import_mode=IMPORT_MODE_REPLACE,
        )

        self.assertEqual(classroom.students.count(), 2)
        self.assertFalse(classroom.students.filter(name="旧学生").exists())
        self.assertEqual(result["created"], 2)
        self.assertEqual(result["updated"], 0)
        self.assertEqual(result["skipped"], 0)


class ClassroomFeatureTests(TestCase):
    def test_search_students_supports_pinyin_initials(self):
        classroom = Classroom.objects.create(name="首字母搜索班", rows=1, cols=2)
        zhangsan = classroom.students.create(name="张三")
        classroom.students.create(name="李四")

        url = reverse("search_students", args=[classroom.pk])
        response = self.client.get(url, {"q": "zs"})

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        student_ids = [item.get("id") for item in payload.get("students", [])]
        self.assertIn(zhangsan.pk, student_ids)

    def test_export_options_pages_render(self):
        classroom = Classroom.objects.create(name="导出配置页", rows=2, cols=2)

        excel_url = reverse("export_students_options_page", args=[classroom.pk])
        svg_url = reverse("export_students_svg_options_page", args=[classroom.pk])
        pptx_url = reverse("export_students_pptx_options_page", args=[classroom.pk])

        excel_resp = self.client.get(excel_url)
        svg_resp = self.client.get(svg_url)
        pptx_resp = self.client.get(pptx_url)

        self.assertEqual(excel_resp.status_code, 200)
        self.assertEqual(svg_resp.status_code, 200)
        self.assertEqual(pptx_resp.status_code, 200)
        self.assertIn("导出排座表", excel_resp.content.decode("utf-8"))
        self.assertIn("导出 SVG 图片", svg_resp.content.decode("utf-8"))
        self.assertIn("导出 PPTX（单页 16:9）", pptx_resp.content.decode("utf-8"))

    def test_import_options_pages_render(self):
        classroom = Classroom.objects.create(name="导入配置页", rows=2, cols=2)

        score_url = reverse("import_students_options_page", args=[classroom.pk])
        layout_url = reverse("import_layout_excel_options_page", args=[classroom.pk])

        score_resp = self.client.get(score_url)
        layout_resp = self.client.get(layout_url)

        self.assertEqual(score_resp.status_code, 200)
        self.assertEqual(layout_resp.status_code, 200)
        self.assertIn("导入 Excel 成绩表", score_resp.content.decode("utf-8"))
        self.assertIn("导入座位表（Excel）", layout_resp.content.decode("utf-8"))
        self.assertIn('id="import-start-row"', score_resp.content.decode("utf-8"))
        self.assertIn('data-import-field="name"', score_resp.content.decode("utf-8"))
        self.assertIn('id="import-preview-area"', score_resp.content.decode("utf-8"))
        self.assertIn('id="student-import-guide-modal"', score_resp.content.decode("utf-8"))

    def test_export_students_default_layout(self):
        classroom = Classroom.objects.create(name="导出默认", rows=2, cols=2)
        seat_a = classroom.seats.get(row=1, col=1)
        seat_d = classroom.seats.get(row=2, col=2)
        seat_a.student = classroom.students.create(name="A")
        seat_d.student = classroom.students.create(name="D")
        seat_a.save(update_fields=["student"])
        seat_d.save(update_fields=["student"])

        url = reverse("export_students", args=[classroom.pk])
        response = self.client.get(url)

        self.assertEqual(response.status_code, 200)
        wb = openpyxl.load_workbook(BytesIO(response.content))
        ws = wb.active
        self.assertEqual(ws.cell(row=2, column=1).value, "讲台")
        self.assertEqual(ws.cell(row=3, column=1).value, "A")
        self.assertEqual(ws.cell(row=4, column=2).value, "D")
        self.assertEqual(ws.cell(row=1, column=1).font.name, "鸿蒙黑体 Medium")
        self.assertEqual(ws.cell(row=3, column=1).font.name, "鸿蒙黑体 Light")

    def test_export_students_rotate_180_layout(self):
        classroom = Classroom.objects.create(name="导出翻转", rows=2, cols=2)
        seat_a = classroom.seats.get(row=1, col=1)
        seat_d = classroom.seats.get(row=2, col=2)
        seat_a.student = classroom.students.create(name="A")
        seat_d.student = classroom.students.create(name="D")
        seat_a.save(update_fields=["student"])
        seat_d.save(update_fields=["student"])

        url = reverse("export_students", args=[classroom.pk]) + "?layout_transform=rotate_180"
        response = self.client.get(url)

        self.assertEqual(response.status_code, 200)
        wb = openpyxl.load_workbook(BytesIO(response.content))
        ws = wb.active
        self.assertIn("180°翻转", ws.cell(row=1, column=1).value or "")
        self.assertEqual(ws.cell(row=4, column=1).value, "讲台")
        self.assertEqual(ws.cell(row=2, column=1).value, "D")
        self.assertEqual(ws.cell(row=3, column=2).value, "A")

    def test_export_students_csis_csls_includes_unknowns_and_seat_extra(self):
        classroom = Classroom.objects.create(name="CSIS班", rows=2, cols=2)
        group = classroom.groups.create(name="第一组", order=1)
        alice = classroom.students.create(name="Alice", student_id="12", gender="M", score=95)
        bob = classroom.students.create(name="Bob")
        seat = classroom.seats.get(row=1, col=2)
        seat.student = alice
        seat.group = group
        seat.save(update_fields=["student", "group"])

        url = reverse("export_students_csis", args=[classroom.pk])
        response = self.client.get(url)

        self.assertEqual(response.status_code, 200)
        self.assertIn("application/json", response.get("Content-Type", ""))
        self.assertIn(".csls", response.get("Content-Disposition", ""))
        payload = json.loads(response.content.decode("utf-8"))
        self.assertEqual(payload["version"], 1)
        csis_class = payload["classes"][0]
        self.assertEqual(csis_class["name"], "CSIS班")
        self.assertEqual({group["name"] for group in csis_class["groups"]}, {"第一组", "unknown"})

        students = {student["name"]: student for student in csis_class["students"]}
        self.assertEqual(students["Alice"]["group"], "第一组")
        self.assertEqual(students["Alice"]["gender"], "male")
        self.assertEqual(students["Alice"]["number"], 12)
        self.assertEqual(students["Alice"]["extra"]["seat"]["row"], 1)
        self.assertEqual(students["Alice"]["extra"]["seat"]["col"], 2)
        self.assertEqual(students["Alice"]["extra"]["seat"]["position"], [1, 0])
        self.assertEqual(students["Alice"]["extra"]["seat"]["coordinate"], "1-2")
        self.assertEqual(students["Bob"]["group"], "unknown")
        self.assertEqual(students["Bob"]["gender"], "unknown")
        self.assertFalse(students["Bob"]["extra"]["seat"]["assigned"])

        detail_response = self.client.get(reverse("classroom_detail", args=[classroom.pk]))
        self.assertContains(detail_response, "导出 CSIS（CSLS）")

    def test_export_students_svg_returns_svg_content(self):
        classroom = Classroom.objects.create(name="SVG班", rows=1, cols=2)
        student = classroom.students.create(name="Alice", score=95)
        seat = classroom.seats.get(row=1, col=1)
        seat.student = student
        seat.save(update_fields=["student"])

        url = reverse("export_students_svg", args=[classroom.pk])
        response = self.client.get(url)

        self.assertEqual(response.status_code, 200)
        self.assertIn("image/svg+xml", response.get("Content-Type", ""))
        content = response.content.decode("utf-8")
        self.assertIn("<svg", content)
        self.assertIn("Alice", content)
        self.assertIn("SVG班", content)
        self.assertNotIn("总座位", content)
        self.assertNotIn("网格", content)
        self.assertIn('font-family:"鸿蒙黑体 Medium"', content)
        self.assertIn('font-family:"鸿蒙黑体 Light"', content)
        self.assertNotIn("PingFang SC", content)
        self.assertNotIn("Microsoft YaHei", content)
        self.assertNotIn("sans-serif", content)

    def test_export_students_svg_respects_visibility_options(self):
        classroom = Classroom.objects.create(name="SVG配置班", rows=1, cols=1)
        student = classroom.students.create(name="Alice", score=95)
        seat = classroom.seats.get(row=1, col=1)
        seat.student = student
        seat.save(update_fields=["student"])

        url = reverse("export_students_svg", args=[classroom.pk]) + (
            "?show_title=0&show_podium=0&show_coords=0&show_name=0&show_score=0&show_empty_label=0"
        )
        response = self.client.get(url)

        self.assertEqual(response.status_code, 200)
        content = response.content.decode("utf-8")
        self.assertNotIn("座次图", content)
        self.assertNotIn("讲台", content)
        self.assertNotIn("Alice", content)
        self.assertNotIn("95分", content)
        self.assertNotIn("(1-1)", content)

    def test_export_students_svg_supports_theme_and_hide_seat_type(self):
        classroom = Classroom.objects.create(name="SVG主题班", rows=1, cols=2)
        seat = classroom.seats.get(row=1, col=2)
        seat.cell_type = SeatCellType.AISLE
        seat.save(update_fields=["cell_type"])

        url = reverse("export_students_svg", args=[classroom.pk]) + "?theme=contrast&show_podium=0&show_seat_type=0"
        response = self.client.get(url)

        self.assertEqual(response.status_code, 200)
        content = response.content.decode("utf-8")
        self.assertIn("#0b1220", content)
        self.assertNotIn("走廊", content)

    def test_export_students_svg_name_with_group_uses_emphasis_layout(self):
        classroom = Classroom.objects.create(name="SVG姓名小组班", rows=1, cols=1)
        group = classroom.groups.create(name="第1组", order=1)
        student = classroom.students.create(name="赵小明")
        seat = classroom.seats.get(row=1, col=1)
        seat.student = student
        seat.group = group
        seat.save(update_fields=["student", "group"])

        url = reverse("export_students_svg", args=[classroom.pk]) + (
            "?show_coords=0&show_name=1&show_score=0&show_group=1"
        )
        response = self.client.get(url)

        self.assertEqual(response.status_code, 200)
        content = response.content.decode("utf-8")
        self.assertIn('class="cell-name" font-size="', content)
        self.assertIn('dominant-baseline="middle"', content)

    @unittest.skipUnless(importlib.util.find_spec("pptx"), "python-pptx not installed")
    def test_export_students_pptx_returns_pptx_content(self):
        classroom = Classroom.objects.create(name="PPT班", rows=1, cols=1)
        student = classroom.students.create(name="Alice", score=91)
        seat = classroom.seats.get(row=1, col=1)
        seat.student = student
        seat.save(update_fields=["student"])

        url = reverse("export_students_pptx", args=[classroom.pk]) + "?show_score=0"
        response = self.client.get(url)

        self.assertEqual(response.status_code, 200)
        self.assertIn(
            "application/vnd.openxmlformats-officedocument.presentationml.presentation",
            response.get("Content-Type", ""),
        )
        self.assertIn(".pptx", response.get("Content-Disposition", ""))
        self.assertTrue(response.content.startswith(b"PK"))
        with zipfile.ZipFile(BytesIO(response.content)) as zf:
            slide_xml = zf.read("ppt/slides/slide1.xml").decode("utf-8")
        self.assertIn('typeface="鸿蒙黑体 Medium"', slide_xml)
        self.assertIn('typeface="鸿蒙黑体 Light"', slide_xml)
        self.assertNotIn('typeface="微软雅黑"', slide_xml)
        self.assertNotIn('typeface="Microsoft YaHei"', slide_xml)

    def test_export_students_svg_preview_student_returns_real_student(self):
        classroom = Classroom.objects.create(name="SVG预览班", rows=1, cols=2)
        s1 = classroom.students.create(name="Alice", score=90)
        s2 = classroom.students.create(name="Bob", score=80)
        seat = classroom.seats.get(row=1, col=1)
        seat.student = s1
        seat.save(update_fields=["student"])

        url = reverse("export_students_svg_preview_student", args=[classroom.pk])
        response = self.client.get(url)

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload.get("status"), "success")
        sample = payload.get("sample") or {}
        self.assertEqual(sample.get("classroom"), "SVG预览班")
        self.assertIn(sample.get("name"), ["Alice", "Bob"])

    def test_export_group_report_header_font_size_code_is_delimited(self):
        classroom = Classroom.objects.create(name="09班", rows=2, cols=2)

        url = reverse("export_group_report", args=[classroom.pk])
        response = self.client.get(url)

        self.assertEqual(response.status_code, 200)
        with zipfile.ZipFile(BytesIO(response.content)) as zf:
            sheet_xml = zf.read("xl/worksheets/sheet1.xml").decode("utf-8")

        self.assertIn("&amp;20 09", sheet_xml)
        self.assertNotIn("&amp;1409", sheet_xml)

    def test_export_group_report_uses_only_harmony_fonts(self):
        classroom = Classroom.objects.create(name="字体班", rows=1, cols=1)
        group = classroom.groups.create(name="一组", order=1)
        student = classroom.students.create(name="张三")
        seat = classroom.seats.get(row=1, col=1)
        seat.student = student
        seat.group = group
        seat.save(update_fields=["student", "group"])

        response = self.client.get(reverse("export_group_report", args=[classroom.pk]))

        self.assertEqual(response.status_code, 200)
        wb = openpyxl.load_workbook(BytesIO(response.content))
        ws = wb.active
        self.assertEqual(ws.cell(row=1, column=1).font.name, "鸿蒙黑体 Medium")
        member_cells = [
            cell
            for row in ws.iter_rows()
            for cell in row
            if cell.value == "张三"
        ]
        self.assertTrue(member_cells)
        self.assertEqual(member_cells[0].font.name, "鸿蒙黑体 Light")
        with zipfile.ZipFile(BytesIO(response.content)) as zf:
            styles_xml = zf.read("xl/styles.xml").decode("utf-8")
            sheet_xml = zf.read("xl/worksheets/sheet1.xml").decode("utf-8")
        self.assertIn('val="鸿蒙黑体 Medium"', styles_xml)
        self.assertIn('val="鸿蒙黑体 Light"', styles_xml)
        self.assertNotIn("微软雅黑", styles_xml + sheet_xml)
        self.assertNotIn("Microsoft YaHei", styles_xml + sheet_xml)


class DesktopExportBridgeTests(TestCase):
    class _FakeDesktopResponse:
        status = 200

        def __enter__(self):
            return self

        def __exit__(self, exc_type, exc, traceback):
            return False

        def read(self):
            return b""

    class _FakeJsonDesktopResponse(_FakeDesktopResponse):
        def read(self):
            return b'{"status":"success","message":"BSCE import ok"}'

    def test_parse_content_disposition_filename_supports_utf8(self):
        value = "attachment; filename*=UTF-8''%E5%BA%A7%E6%AC%A1%E5%9B%BE.svg"
        self.assertEqual(parse_content_disposition_filename(value), "座次图.svg")

    def test_resolve_local_export_url_rejects_external_origin(self):
        with self.assertRaises(ValueError):
            resolve_local_export_url("http://127.0.0.1:23948", "https://example.com/test.xlsx")

    def test_normalize_accept_extensions_falls_back_to_filename_suffix(self):
        self.assertEqual(normalize_accept_extensions([], "示例.pptx"), [".pptx"])

    def test_ensure_allowed_extension_preserves_known_suffix(self):
        self.assertEqual(
            ensure_allowed_extension("/tmp/export.json", [".seats", ".json"]),
            "/tmp/export.json",
        )

    def test_is_allowed_extension_accepts_seats_snapshot(self):
        self.assertTrue(is_allowed_extension("/tmp/测试班.seats", [".seats", ".json"]))
        self.assertFalse(is_allowed_extension("/tmp/1000090715.jpg", [".seats", ".json"]))

    def test_build_file_dialog_types_uses_expected_mask(self):
        file_types = build_file_dialog_types(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            [".xlsx"],
        )
        self.assertEqual(file_types[0], "Excel 文件 (*.xlsx)")

    def test_build_multipart_form_data_contains_seats_file(self):
        boundary, body = build_multipart_form_data(
            fields={"csrfmiddlewaretoken": "csrf-token"},
            files=[
                {
                    "field_name": "seats_file",
                    "filename": "测试班.seats",
                    "content_type": "application/octet-stream",
                    "content": b'{"current_state":{}}',
                }
            ],
        )

        self.assertTrue(boundary.startswith("----FuckSeatsBoundary"))
        self.assertIn(b'name="csrfmiddlewaretoken"', body)
        self.assertIn("filename=\"测试班.seats\"".encode("utf-8"), body)
        self.assertIn(b'{"current_state":{}}', body)

    def test_desktop_import_seats_file_uses_unfiltered_open_dialog(self):
        class FakeWindow:
            def __init__(self, selected_path):
                self.selected_path = selected_path
                self.calls = []

            def create_file_dialog(self, dialog_type, **kwargs):
                self.calls.append((dialog_type, kwargs))
                return [self.selected_path]

        with tempfile.TemporaryDirectory() as temp_dir:
            seats_path = Path(temp_dir) / "测试班.seats"
            seats_path.write_text('{"current_state":{}}', encoding="utf-8")
            fake_window = FakeWindow(str(seats_path))
            bridge = DesktopBridge("http://127.0.0.1:23948")
            bridge._attach_window(fake_window)
            captured_requests = []

            def fake_urlopen(req, timeout=None):
                captured_requests.append(req)
                return self._FakeDesktopResponse()

            with patch("desktop_shell.urllib.request.urlopen", side_effect=fake_urlopen):
                result = bridge.import_seats_file(
                    "/classroom/1/layout/import/",
                    "csrf-token",
                    [".seats", ".json"],
                )

        self.assertEqual(result["status"], "imported")
        self.assertEqual(fake_window.calls[0][1]["file_types"], ())
        self.assertIn("测试班.seats".encode("utf-8"), captured_requests[0].data)
        self.assertEqual(captured_requests[0].get_header("X-csrftoken"), "csrf-token")
        self.assertIn("csrftoken=csrf-token", captured_requests[0].get_header("Cookie"))

    def test_desktop_upload_local_file_supports_custom_extension_without_dialog_filter(self):
        class FakeWindow:
            def __init__(self, selected_path):
                self.selected_path = selected_path
                self.calls = []

            def create_file_dialog(self, dialog_type, **kwargs):
                self.calls.append((dialog_type, kwargs))
                return [self.selected_path]

        with tempfile.TemporaryDirectory() as temp_dir:
            sce_path = Path(temp_dir) / "座位表.sce"
            sce_path.write_text('{"schema":"bsce"}', encoding="utf-8")
            fake_window = FakeWindow(str(sce_path))
            bridge = DesktopBridge("http://127.0.0.1:23948")
            bridge._attach_window(fake_window)
            captured_requests = []

            def fake_urlopen(req, timeout=None):
                captured_requests.append(req)
                return self._FakeJsonDesktopResponse()

            with patch("desktop_shell.urllib.request.urlopen", side_effect=fake_urlopen):
                result = bridge.upload_local_file(
                    "/classroom/1/layout/import/bsce/",
                    "csrf-token",
                    "bsce_file",
                    [".sce"],
                )

        self.assertEqual(result["status"], "success")
        self.assertEqual(result["message"], "BSCE import ok")
        self.assertEqual(fake_window.calls[0][1]["file_types"], ())
        self.assertIn(b'name="bsce_file"', captured_requests[0].data)
        self.assertEqual(captured_requests[0].get_header("X-requested-with"), "XMLHttpRequest")


class FrontendStoreTests(TestCase):
    def test_frontend_store_js_embeds_backend_store_and_runtime_patch(self):
        FrontendKVStore.objects.create(key="seats_plugin_dev_mode", value="1")

        response = self.client.get(reverse("frontend_store_js"))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/javascript")
        content = response.content.decode("utf-8")
        self.assertIn('"seats_plugin_dev_mode": "1"', content)
        self.assertIn("Storage.prototype.setItem", content)
        self.assertIn("syncBackendSet", content)
        self.assertIn("window.BACKEND_STORE[key] === value", content)
        self.assertIn("mac_os_warning_seen", content)

    def test_frontend_store_set_upserts_value(self):
        url = reverse("frontend_store_set")

        response = self.client.post(
            url,
            data=json.dumps({"key": "plugin_mode", "value": "teacher"}),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["status"], "success")
        self.assertEqual(FrontendKVStore.objects.get(key="plugin_mode").value, "teacher")

        response = self.client.post(
            url,
            data=json.dumps({"key": "plugin_mode", "value": "dev"}),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(FrontendKVStore.objects.count(), 1)
        self.assertEqual(FrontendKVStore.objects.get(key="plugin_mode").value, "dev")

    def test_frontend_store_set_skips_invalid_key_without_bad_request(self):
        response = self.client.post(
            reverse("frontend_store_set"),
            data=json.dumps({"key": "x" * 256, "value": "dev"}),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["status"], "success")
        self.assertFalse(response.json()["persisted"])
        self.assertEqual(FrontendKVStore.objects.count(), 0)

    def test_frontend_store_set_stringifies_supported_values(self):
        response = self.client.post(
            reverse("frontend_store_set"),
            data=json.dumps({"key": 0, "value": None}),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()["persisted"])
        self.assertEqual(FrontendKVStore.objects.get(key="0").value, "")

    def test_frontend_store_delete_removes_value(self):
        FrontendKVStore.objects.create(key="plugin_mode", value="dev")

        response = self.client.post(
            reverse("frontend_store_delete"),
            data=json.dumps({"key": "plugin_mode"}),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["status"], "success")
        self.assertFalse(FrontendKVStore.objects.filter(key="plugin_mode").exists())

    def test_cloud_config_persists_data_sharing_prompt_decision(self):
        with patch("desktop_runtime.get_current_version", return_value="9.8.7"):
            response = self.client.post(
                reverse("cloud_config"),
                data=json.dumps({
                    "data_sharing_enabled": True,
                    "data_sharing_prompt_seen_version": "9.8.7",
                }),
                content_type="application/json",
            )

            self.assertEqual(response.status_code, 200)
            self.assertEqual(response.json()["status"], "success")
            self.assertEqual(
                FrontendKVStore.objects.get(key=DATA_SHARING_PROMPT_SEEN_VERSION_KEY).value,
                "9.8.7",
            )
            self.assertEqual(
                FrontendKVStore.objects.get(key=DATA_SHARING_ENABLED_KEY).value,
                "1",
            )
            self.assertFalse(response.json()["data_sharing"]["show_prompt"])

    def test_mark_onboarding_seen_persists_stable_frontend_store_key(self):
        response = self.client.post(
            reverse("mark_onboarding_seen"),
            data=json.dumps({"completed_steps": "detail_done"}),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["ok"], True)
        self.assertEqual(
            FrontendKVStore.objects.get(key=ONBOARDING_SEEN_STORE_KEY).value,
            ONBOARDING_SEEN_STORE_VALUE,
        )
        state = OnboardingState.objects.get()
        self.assertTrue(state.seen)
        self.assertEqual(state.completed_steps, "detail_done")
        self.assertFalse(response.json()["sample_deleted"])

    def test_index_creates_onboarding_sample_with_groups(self):
        response = self.client.get(reverse("index"))

        self.assertEqual(response.status_code, 200)
        sample = Classroom.objects.get(name=ONBOARDING_SAMPLE_NAME)
        self.assertEqual(sample.students.count(), 12)
        self.assertEqual(list(sample.groups.values_list("name", flat=True)), ["第一组", "第二组"])

    def test_mark_onboarding_seen_deletes_completed_sample_classroom(self):
        sample = Classroom.objects.create(name=ONBOARDING_SAMPLE_NAME, rows=2, cols=2)
        session = self.client.session
        session["onboarding_sample_pk"] = sample.pk
        session.save()

        response = self.client.post(
            reverse("mark_onboarding_seen"),
            data=json.dumps({
                "completed_steps": "detail_done",
                "current_classroom_id": sample.pk,
            }),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["ok"])
        self.assertFalse(payload["sample_deleted"])
        self.assertNotIn("redirect_url", payload)
        self.assertTrue(Classroom.objects.filter(pk=sample.pk).exists())

        response = self.client.get(reverse("index"))
        self.assertEqual(response.status_code, 200)
        self.assertFalse(Classroom.objects.filter(pk=sample.pk).exists())

    def test_index_uses_stable_frontend_store_key_to_suppress_onboarding(self):
        FrontendKVStore.objects.create(
            key=ONBOARDING_SEEN_STORE_KEY,
            value=ONBOARDING_SEEN_STORE_VALUE,
        )

        response = self.client.get(reverse("index"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "window.ONBOARDING_SHOULD_SHOW = false;")
        self.assertFalse(Classroom.objects.filter(name="示例班级（新手引导）").exists())

    def test_index_backfills_stable_key_from_legacy_onboarding_state(self):
        OnboardingState.objects.create(
            session_key="old-session",
            seen=True,
            completed_steps="detail_done",
        )

        response = self.client.get(reverse("index"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "window.ONBOARDING_SHOULD_SHOW = false;")
        self.assertEqual(
            FrontendKVStore.objects.get(key=ONBOARDING_SEEN_STORE_KEY).value,
            ONBOARDING_SEEN_STORE_VALUE,
        )
        self.assertFalse(Classroom.objects.filter(name="示例班级（新手引导）").exists())

    @override_settings(APP_SHELL="webview")
    def test_index_template_exposes_app_shell_runtime(self):
        response = self.client.get(reverse("index"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'data-app-shell="webview"')

    def test_rename_classroom_success(self):
        classroom = Classroom.objects.create(name="原班级", rows=2, cols=2)
        url = reverse("rename_classroom", args=[classroom.pk])

        response = self.client.post(
            url,
            data=json.dumps({"name": "新班级名称"}),
            content_type="application/json",
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json().get("status"), "success")
        classroom.refresh_from_db()
        self.assertEqual(classroom.name, "新班级名称")

    def test_rename_classroom_rejects_empty_name(self):
        classroom = Classroom.objects.create(name="原班级2", rows=2, cols=2)
        url = reverse("rename_classroom", args=[classroom.pk])

        response = self.client.post(
            url,
            data=json.dumps({"name": "   "}),
            content_type="application/json",
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.json().get("status"), "error")
        classroom.refresh_from_db()
        self.assertEqual(classroom.name, "原班级2")


class CloudSyncHeaderTests(TestCase):
    class _FakeJsonResponse:
        status = 200

        def __enter__(self):
            return self

        def __exit__(self, exc_type, exc, traceback):
            return False

        def read(self):
            return b'{"status":"success"}'

    def test_cloud_request_uses_ssl_context(self):
        context = ssl.create_default_context()
        captured_calls = []

        def fake_urlopen(req, timeout=None, context=None):
            captured_calls.append((req, timeout, context))
            return self._FakeJsonResponse()

        with patch("seats.cloud._get_ssl_context", return_value=context):
            with patch("seats.cloud.urllib.request.urlopen", side_effect=fake_urlopen):
                payload = cloud_module._request_json(
                    "POST",
                    "https://example.test/api",
                    {"developer": "老三"},
                    timeout=9,
                )

        self.assertEqual(payload["status"], "success")
        self.assertEqual(payload["_http_status"], 200)
        self.assertEqual(captured_calls[0][1], 9)
        self.assertIs(captured_calls[0][2], context)
        self.assertEqual(captured_calls[0][0].get_header("User-agent"), "fuckseats_cilent")

    def test_cloud_ssl_context_builds_ssl_context(self):
        self.assertIsInstance(cloud_module._get_ssl_context(), ssl.SSLContext)

    def create_cloud_session(self):
        return CloudSession.objects.create(
            uid="u1",
            nickname="老三",
            session_token="token",
            token_expires_at=timezone.now() + timedelta(days=1),
        )

    def cloud_classroom_payload(self, name="云端班级", rows=3, cols=3):
        return {
            "current_state": {
                "classroom": {
                    "pk": 9001,
                    "name": name,
                    "rows": rows,
                    "cols": cols,
                    "left_guardian_student_pk": None,
                    "right_guardian_student_pk": None,
                    "created_at": "",
                },
                "students": [],
                "groups": [],
                "seats": [],
                "constraints": [],
                "layout_snapshots": [],
            },
            "history": {"entries": []},
            "future_mode_config": None,
            "ai_conversations": [],
        }

    def test_classroom_detail_embeds_sync_pill_runtime(self):
        classroom = Classroom.objects.create(name="同步班", rows=2, cols=2)

        response = self.client.get(reverse("classroom_detail", args=[classroom.pk]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "classroom-sync-meta")
        self.assertContains(response, "cloud-sync-pill")
        self.assertContains(response, "FuckSeatsCloudSync")
        self.assertContains(response, "点击上云")

    def test_classroom_state_includes_sync_meta(self):
        classroom = Classroom.objects.create(name="状态班", rows=2, cols=2)
        meta = SyncMeta.objects.get(classroom=classroom)
        meta.cloud_version = 2
        meta.local_version = 3
        meta.last_sync_at = timezone.now()
        meta.save(update_fields=["cloud_version", "local_version", "last_sync_at", "updated_at"])

        response = self.client.get(reverse("classroom_state", args=[classroom.pk]))

        self.assertEqual(response.status_code, 200)
        payload = response.json()["sync_meta"]
        self.assertTrue(payload["backed_up"])
        self.assertTrue(payload["has_local_changes"])
        self.assertEqual(payload["state"], "dirty")

    def test_cloud_userinfo_refreshes_subscription_when_logged_in(self):
        self.create_cloud_session()

        with patch("seats.views.refresh_cloud_subscription") as refresh_subscription:
            def fake_refresh(session, strict=False, timeout=20):
                session.subscription_tier = "pro"
                session.subscription_display_name = "专业版"
                session.limits = {"max_classrooms": 8}
                return session

            refresh_subscription.side_effect = fake_refresh
            response = self.client.get(reverse("cloud_userinfo"))

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["logged_in"])
        self.assertEqual(payload["tier"], "pro")
        self.assertEqual(payload["tier_display"], "专业版")
        self.assertEqual(payload["limits"]["max_classrooms"], 8)
        refresh_subscription.assert_called_once()

    def test_cloud_sync_skips_up_to_date_classroom(self):
        classroom = Classroom.objects.create(name="云端最新班", rows=2, cols=2)
        meta = SyncMeta.objects.get(classroom=classroom)
        meta.cloud_version = 4
        meta.local_version = 4
        meta.last_sync_at = timezone.now()
        meta.save(update_fields=["cloud_version", "local_version", "last_sync_at", "updated_at"])
        self.create_cloud_session()

        with patch("seats.views.cloud_api_request") as cloud_request:
            cloud_request.return_value = {"ok": True, "status": "success", "versions": {str(meta.uuid): 4}}
            response = self.client.post(
                reverse("cloud_sync"),
                data=json.dumps({"classroom_ids": [classroom.pk]}),
                content_type="application/json",
            )

        self.assertEqual(response.status_code, 200)
        row = response.json()["results"][0]
        self.assertEqual(row["status"], "up_to_date")
        self.assertEqual(row["state"], "synced")
        cloud_request.assert_called_once()
        self.assertEqual(cloud_request.call_args.args[1:3], ("GET", "/api/sync/status"))

    def test_cloud_sync_uses_refreshed_subscription_limits_before_sync(self):
        classroom_a = Classroom.objects.create(name="班级A", rows=2, cols=2)
        classroom_b = Classroom.objects.create(name="班级B", rows=2, cols=2)
        self.create_cloud_session()

        with patch("seats.views.refresh_cloud_subscription") as refresh_subscription:
            with patch("seats.views.cloud_api_request") as cloud_request:
                def fake_refresh(session, strict=False, timeout=20):
                    session.subscription_tier = "pro"
                    session.subscription_display_name = "专业版"
                    session.limits = {"max_classrooms": 1}
                    return session

                refresh_subscription.side_effect = fake_refresh
                cloud_request.side_effect = [
                    {"ok": True, "status": "success", "versions": {}, "classrooms": []},
                    {"ok": True, "status": "success", "uuid": str(SyncMeta.objects.get(classroom=classroom_a).uuid), "version": 1},
                ]
                response = self.client.post(
                    reverse("cloud_sync"),
                    data=json.dumps({"classroom_ids": [classroom_a.pk, classroom_b.pk]}),
                    content_type="application/json",
                )

        self.assertEqual(response.status_code, 200)
        results = response.json()["results"]
        self.assertEqual(results[0]["status"], "ok")
        self.assertEqual(results[1]["status"], "skipped")
        self.assertIn("最多同步 1 个班级", results[1]["message"])
        refresh_subscription.assert_called_once()

    def test_cloud_sync_pushes_when_remote_record_is_missing(self):
        classroom = Classroom.objects.create(name="云端缺失班", rows=2, cols=2)
        meta = SyncMeta.objects.get(classroom=classroom)
        meta.cloud_version = 4
        meta.local_version = 4
        meta.last_sync_at = timezone.now()
        meta.save(update_fields=["cloud_version", "local_version", "last_sync_at", "updated_at"])
        self.create_cloud_session()

        with patch("seats.views.cloud_api_request") as cloud_request:
            cloud_request.side_effect = [
                {"ok": True, "status": "success", "versions": {}},
                {"ok": True, "status": "success", "uuid": str(meta.uuid), "version": 1},
            ]
            response = self.client.post(
                reverse("cloud_sync"),
                data=json.dumps({"classroom_ids": [classroom.pk]}),
                content_type="application/json",
            )

        self.assertEqual(response.status_code, 200)
        row = response.json()["results"][0]
        self.assertEqual(row["status"], "ok")
        self.assertEqual(row["cloud_version"], 1)
        self.assertEqual(cloud_request.call_count, 2)
        push_body = cloud_request.call_args_list[1].args[3]
        self.assertEqual(push_body["base_version"], 4)
        self.assertEqual(push_body["cloud_version"], 4)

    def test_cloud_sync_reports_error_when_remote_status_unavailable(self):
        classroom = Classroom.objects.create(name="云端不可用班", rows=2, cols=2)
        self.create_cloud_session()

        with patch("seats.views.cloud_api_request") as cloud_request:
            cloud_request.side_effect = CloudAPIError("无法连接云端服务：拒绝连接", status_code=502)
            response = self.client.post(
                reverse("cloud_sync"),
                data=json.dumps({"classroom_ids": [classroom.pk]}),
                content_type="application/json",
            )

        self.assertEqual(response.status_code, 502)
        self.assertIn("无法连接云端服务", response.json()["message"])

    def test_cloud_sync_conflicts_when_remote_version_is_newer(self):
        classroom = Classroom.objects.create(name="远端更新班", rows=2, cols=2)
        meta = SyncMeta.objects.get(classroom=classroom)
        meta.cloud_version = 4
        meta.local_version = 6
        meta.last_sync_at = timezone.now()
        meta.save(update_fields=["cloud_version", "local_version", "last_sync_at", "updated_at"])
        self.create_cloud_session()

        with patch("seats.views.cloud_api_request") as cloud_request:
            cloud_request.return_value = {"ok": True, "status": "success", "versions": {str(meta.uuid): 5}}
            response = self.client.post(
                reverse("cloud_sync"),
                data=json.dumps({"classroom_ids": [classroom.pk]}),
                content_type="application/json",
            )

        self.assertEqual(response.status_code, 200)
        row = response.json()["results"][0]
        self.assertEqual(row["status"], "conflict")
        self.assertEqual(row["cloud_version"], 5)
        cloud_request.assert_called_once()

    def test_cloud_sync_pulls_when_remote_is_newer_without_local_changes(self):
        classroom = Classroom.objects.create(name="本地旧班", rows=2, cols=2)
        meta = SyncMeta.objects.get(classroom=classroom)
        meta.cloud_version = 4
        meta.local_version = 4
        meta.last_sync_at = timezone.now()
        meta.save(update_fields=["cloud_version", "local_version", "last_sync_at", "updated_at"])
        self.create_cloud_session()

        with patch("seats.views.cloud_api_request") as cloud_request:
            cloud_request.side_effect = [
                {"ok": True, "status": "success", "versions": {str(meta.uuid): 5}, "classrooms": [
                    {"uuid": str(meta.uuid), "name": "云端新班", "version": 5}
                ]},
                {"ok": True, "status": "success", "uuid": str(meta.uuid), "version": 5, "data": self.cloud_classroom_payload("云端新班", 3, 3)},
            ]
            response = self.client.post(
                reverse("cloud_sync"),
                data=json.dumps({"classroom_ids": [classroom.pk]}),
                content_type="application/json",
            )

        self.assertEqual(response.status_code, 200)
        row = response.json()["results"][0]
        self.assertEqual(row["status"], "pulled")
        self.assertEqual(row["cloud_version"], 5)
        classroom.refresh_from_db()
        self.assertEqual(classroom.name, "云端新班")
        self.assertEqual(classroom.rows, 3)
        self.assertEqual(cloud_request.call_count, 2)
        self.assertEqual(cloud_request.call_args_list[1].args[1:3], ("GET", f"/api/sync/pull/{meta.uuid}"))

    def test_cloud_sync_pulls_remote_only_classrooms_on_full_sync(self):
        remote_uuid = uuid.uuid4()
        self.create_cloud_session()

        with patch("seats.views.cloud_api_request") as cloud_request:
            cloud_request.side_effect = [
                {"ok": True, "status": "success", "versions": {str(remote_uuid): 3}, "classrooms": [
                    {"uuid": str(remote_uuid), "name": "只在云端的班", "version": 3}
                ]},
                {"ok": True, "status": "success", "uuid": str(remote_uuid), "version": 3, "data": self.cloud_classroom_payload("只在云端的班", 4, 5)},
            ]
            response = self.client.post(
                reverse("cloud_sync"),
                data=json.dumps({}),
                content_type="application/json",
            )

        self.assertEqual(response.status_code, 200)
        row = response.json()["results"][0]
        self.assertEqual(row["status"], "pulled")
        self.assertTrue(row["remote_only"])
        classroom = Classroom.objects.get(name="只在云端的班")
        meta = SyncMeta.objects.get(classroom=classroom)
        self.assertEqual(str(meta.uuid), str(remote_uuid))
        self.assertEqual(meta.cloud_version, 3)
        self.assertEqual(classroom.rows, 4)
        self.assertEqual(classroom.cols, 5)

    def test_cloud_sync_force_pushes_when_remote_version_is_newer(self):
        classroom = Classroom.objects.create(name="强制更新班", rows=2, cols=2)
        meta = SyncMeta.objects.get(classroom=classroom)
        meta.cloud_version = 4
        meta.local_version = 4
        meta.last_sync_at = timezone.now()
        meta.save(update_fields=["cloud_version", "local_version", "last_sync_at", "updated_at"])
        self.create_cloud_session()

        with patch("seats.views.cloud_api_request") as cloud_request:
            cloud_request.side_effect = [
                {"ok": True, "status": "success", "versions": {str(meta.uuid): 5}},
                {"ok": True, "status": "success", "uuid": str(meta.uuid), "version": 6},
            ]
            response = self.client.post(
                reverse("cloud_sync"),
                data=json.dumps({"classroom_ids": [classroom.pk], "force": True}),
                content_type="application/json",
            )

        self.assertEqual(response.status_code, 200)
        row = response.json()["results"][0]
        self.assertEqual(row["status"], "ok")
        self.assertEqual(row["cloud_version"], 6)
        self.assertEqual(cloud_request.call_count, 2)
        push_body = cloud_request.call_args_list[1].args[3]
        self.assertIs(push_body["force"], True)
        self.assertEqual(push_body["base_version"], 4)

    def test_cloud_callback_runs_auto_sync_after_login(self):
        classroom = Classroom.objects.create(name="登录同步班", rows=2, cols=2)
        meta = SyncMeta.objects.get(classroom=classroom)
        token_expires_at = timezone.now() + timedelta(days=1)

        with patch("seats.views.cloud_exchange_session_code") as exchange_code:
            with patch("seats.views.cloud_api_request") as cloud_request:
                exchange_code.return_value = {
                    "status": "success",
                    "session_token": "session-token",
                    "token_expires_at": token_expires_at.isoformat(),
                    "uid": "u-login",
                    "nickname": "老三",
                    "subscription": {
                        "tier": "free",
                        "display_name": "免费版",
                        "limits": {"max_classrooms": 3},
                    },
                }
                cloud_request.side_effect = [
                    {"ok": True, "status": "success", "versions": {}, "classrooms": []},
                    {"ok": True, "status": "success", "uuid": str(meta.uuid), "version": 1},
                ]

                response = self.client.get(reverse("cloud_callback") + "?code=login-code")

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response["Location"], "/?cloud_login=success")
        self.assertEqual(cloud_request.call_count, 2)
        push_body = cloud_request.call_args_list[1].args[3]
        self.assertEqual(push_body["uuid"], str(meta.uuid))
        self.assertEqual(push_body["device_id"], "login-auto-sync")
        meta.refresh_from_db()
        self.assertEqual(meta.cloud_version, 1)
        self.assertEqual(meta.local_version, 1)


class PluginSystemTests(TestCase):
    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        self.addCleanup(self.temp_dir.cleanup)
        self.addCleanup(plugin_registry.reset_for_tests)

        plugin_code = """
PLUGIN_META = {
    'id': 'test_plugin',
    'name': '测试插件',
    'version': '1.0.0',
}

HOOK_EVENTS = []


def _on_classroom_created(context):
    classroom = context.get('classroom')
    HOOK_EVENTS.append({
        'event': 'classroom_created',
        'classroom_id': classroom.pk if classroom else None,
    })


def _echo(context):
    payload = context.get('payload') or {}
    classroom = context.get('classroom')
    return {
        'echo': payload,
        'classroom_id': classroom.pk if classroom else None,
    }


def _hook_count(context):
    return {
        'count': len(HOOK_EVENTS),
        'events': HOOK_EVENTS,
    }


UI_SCRIPT = "count = classroom.students.count() if classroom else 0\\nui = components.page(title='测试UI', blocks=[components.metric('学生人数', count)])"
WORKSPACE_SCRIPT = "const marker = document.createElement('div'); marker.id = 'plugin-workspace-test'; document.body.appendChild(marker); return () => marker.remove();"


def register(registry):
    registry.register_hook('classroom_created', _on_classroom_created)
    registry.register_action('echo', _echo, methods=('POST',))
    registry.register_action('hook_count', _hook_count, methods=('GET',))
    registry.register_ui_script('dashboard', UI_SCRIPT, methods=('GET', 'POST'))
    registry.register_workspace_script('inject_marker', WORKSPACE_SCRIPT, methods=('GET',), requires_permission=True, auto_run=True)
""".strip()

        plugin_file = Path(self.temp_dir.name) / 'test_plugin.py'
        plugin_file.write_text(plugin_code, encoding='utf-8')

    @override_settings(PLUGIN_DIRS=[])
    def test_plugins_overview_and_action_dispatch(self):
        with self.settings(PLUGIN_DIRS=[self.temp_dir.name]):
            plugin_registry.reset_for_tests()

            overview_url = reverse('plugins_overview')
            overview_resp = self.client.get(overview_url)
            self.assertEqual(overview_resp.status_code, 200)
            payload = overview_resp.json()
            self.assertEqual(payload.get('status'), 'success')

            components_resp = self.client.get(reverse('plugin_components_overview'))
            self.assertEqual(components_resp.status_code, 200)
            components_payload = components_resp.json()
            self.assertEqual(components_payload.get('status'), 'success')
            self.assertIn('metric', components_payload.get('components') or [])

            plugins = payload.get('plugins') or []
            plugin_ids = [item.get('id') for item in plugins]
            self.assertIn('test_plugin', plugin_ids)
            plugin_row = next(item for item in plugins if item.get('id') == 'test_plugin')
            ui_scripts = plugin_row.get('ui_scripts') or []
            ui_script_names = [item.get('name') for item in ui_scripts]
            self.assertIn('dashboard', ui_script_names)

            create_url = reverse('create_classroom')
            create_resp = self.client.post(create_url, {'name': '插件班级', 'rows': 2, 'cols': 2})
            self.assertEqual(create_resp.status_code, 302)

            classroom = Classroom.objects.get(name='插件班级')
            hook_resp = self.client.get(reverse('plugin_action_dispatch', args=['test_plugin', 'hook_count']))
            self.assertEqual(hook_resp.status_code, 200)
            hook_payload = hook_resp.json().get('result') or {}
            self.assertEqual(hook_payload.get('count'), 1)
            self.assertEqual((hook_payload.get('events') or [])[0].get('classroom_id'), classroom.pk)

            echo_resp = self.client.post(
                reverse('plugin_action_dispatch', args=['test_plugin', 'echo']),
                data=json.dumps({'classroom_id': classroom.pk, 'hello': 'world'}),
                content_type='application/json',
            )
            self.assertEqual(echo_resp.status_code, 200)
            echo_payload = echo_resp.json().get('result') or {}
            self.assertEqual(echo_payload.get('classroom_id'), classroom.pk)
            self.assertEqual((echo_payload.get('echo') or {}).get('hello'), 'world')

            ui_resp = self.client.get(
                reverse('plugin_ui_dispatch', args=['test_plugin', 'dashboard']) + f'?classroom_id={classroom.pk}'
            )
            self.assertEqual(ui_resp.status_code, 200)
            ui_payload = ui_resp.json().get('ui') or {}
            self.assertEqual(ui_payload.get('type'), 'page')
            self.assertEqual(ui_payload.get('title'), '测试UI')
            blocks = ui_payload.get('blocks') or []
            self.assertTrue(blocks)
            self.assertEqual(blocks[0].get('type'), 'metric')
            self.assertEqual(blocks[0].get('value'), 0)

            ui_page_resp = self.client.get(
                reverse('plugin_ui_page', args=['test_plugin', 'dashboard']) + f'?classroom_id={classroom.pk}'
            )
            self.assertEqual(ui_page_resp.status_code, 200)
            page_html = ui_page_resp.content.decode('utf-8')
            self.assertIn('id="plugin-ui-root"', page_html)
            self.assertIn('data-plugin-id="test_plugin"', page_html)

            ext_page_resp = self.client.get(reverse('extensions_overview'))
            self.assertEqual(ext_page_resp.status_code, 200)
            self.assertContains(ext_page_resp, '扩展清单')

            ext_list_resp = self.client.get(
                reverse('extensions_overview') + f'?classroom_id={classroom.pk}',
                HTTP_X_REQUESTED_WITH='XMLHttpRequest',
            )
            self.assertEqual(ext_list_resp.status_code, 200)
            ext_payload = ext_list_resp.json()
            self.assertEqual(ext_payload.get('status'), 'success')
            ext_rows = ext_payload.get('extensions') or []
            ext_ids = [item.get('id') for item in ext_rows]
            self.assertIn('test_plugin', ext_ids)
            test_ext_row = next(item for item in ext_rows if item.get('id') == 'test_plugin')
            self.assertTrue(test_ext_row.get('workspace_permission_required'))
            self.assertFalse(test_ext_row.get('workspace_permission_granted'))

            manifest_resp = self.client.get(reverse('extension_manifest', args=['test_plugin']))
            self.assertEqual(manifest_resp.status_code, 200)
            manifest_payload = manifest_resp.json()
            self.assertEqual(manifest_payload.get('manifest_version'), 3)
            self.assertEqual(manifest_payload.get('short_name'), 'test_plugin')
            self.assertIn('components_library', manifest_payload.get('endpoints') or {})

            runtime_action_resp = self.client.post(
                reverse('extension_send_message', args=['test_plugin']),
                data=json.dumps({
                    'classroom_id': classroom.pk,
                    'message': {
                        'type': 'action',
                        'name': 'echo',
                        'method': 'POST',
                        'payload': {'k': 'v'},
                    }
                }),
                content_type='application/json',
            )
            self.assertEqual(runtime_action_resp.status_code, 200)
            runtime_action_payload = runtime_action_resp.json()
            self.assertEqual(runtime_action_payload.get('status'), 'success')
            runtime_action_result = runtime_action_payload.get('result') or {}
            self.assertEqual((runtime_action_result.get('echo') or {}).get('k'), 'v')

            runtime_ui_resp = self.client.post(
                reverse('extension_send_message', args=['test_plugin']),
                data=json.dumps({
                    'classroom_id': classroom.pk,
                    'message': {
                        'type': 'ui',
                        'name': 'dashboard',
                        'method': 'GET',
                    }
                }),
                content_type='application/json',
            )
            self.assertEqual(runtime_ui_resp.status_code, 200)
            runtime_ui_payload = runtime_ui_resp.json()
            self.assertEqual(runtime_ui_payload.get('status'), 'success')
            self.assertEqual((runtime_ui_payload.get('result') or {}).get('type'), 'page')

            runtime_workspace_forbidden = self.client.post(
                reverse('extension_send_message', args=['test_plugin']),
                data=json.dumps({
                    'classroom_id': classroom.pk,
                    'message': {
                        'type': 'workspace_script',
                        'name': 'inject_marker',
                        'method': 'GET',
                    }
                }),
                content_type='application/json',
            )
            self.assertEqual(runtime_workspace_forbidden.status_code, 403)

            permission_grant_resp = self.client.post(
                reverse('extension_workspace_permission', args=['test_plugin']),
                data=json.dumps({'classroom_id': classroom.pk, 'granted': True}),
                content_type='application/json',
            )
            self.assertEqual(permission_grant_resp.status_code, 200)
            self.assertTrue(permission_grant_resp.json().get('granted'))

            runtime_workspace_resp = self.client.post(
                reverse('extension_send_message', args=['test_plugin']),
                data=json.dumps({
                    'classroom_id': classroom.pk,
                    'message': {
                        'type': 'workspace_script',
                        'name': 'inject_marker',
                        'method': 'GET',
                    }
                }),
                content_type='application/json',
            )
            self.assertEqual(runtime_workspace_resp.status_code, 200)
            workspace_payload = runtime_workspace_resp.json()
            self.assertEqual(workspace_payload.get('status'), 'success')
            self.assertEqual((workspace_payload.get('result') or {}).get('name'), 'inject_marker')
            self.assertIn('source', workspace_payload.get('result') or {})

            permission_query_resp = self.client.get(
                reverse('extension_workspace_permission', args=['test_plugin']) + f'?classroom_id={classroom.pk}'
            )
            self.assertEqual(permission_query_resp.status_code, 200)
            self.assertTrue(permission_query_resp.json().get('granted'))

            workspace_resp = self.client.get(reverse('classroom_detail', args=[classroom.pk]))
            self.assertEqual(workspace_resp.status_code, 200)
            workspace_html = workspace_resp.content.decode('utf-8')
            self.assertIn('id="btn-open-plugin-hub"', workspace_html)
            self.assertIn('data-extensions-list-url="/extensions/"', workspace_html)


class ClassroomCommandApiTests(TestCase):
    def setUp(self):
        self.classroom = Classroom.objects.create(name='命令测试班', rows=3, cols=3)
        self.command_url = reverse('classroom_command', args=[self.classroom.pk])
        self.ai_chat_url = reverse('ai_chat', args=[self.classroom.pk])
        self.ai_chat_stream_url = reverse('ai_chat_stream', args=[self.classroom.pk])

        self.student_a = Student.objects.create(classroom=self.classroom, name='张三', student_id='001', score=98)
        self.student_b = Student.objects.create(classroom=self.classroom, name='李四', student_id='002', score=93)
        self.student_c = Student.objects.create(classroom=self.classroom, name='王五', student_id='003', score=88)

        self.seat_a = self.classroom.seats.get(row=1, col=1)
        self.seat_b = self.classroom.seats.get(row=1, col=2)
        self.seat_a.student = self.student_a
        self.seat_a.save(update_fields=['student'])
        self.seat_b.student = self.student_b
        self.seat_b.save(update_fields=['student'])

    def _post_command(self, command_text):
        return self.client.post(
            self.command_url,
            data=json.dumps({'command': command_text}),
            content_type='application/json',
        )

    def test_classroom_command_manifest_can_be_fetched(self):
        response = self.client.get(self.command_url)
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload.get('status'), 'success')
        manifest = payload.get('manifest') or {}
        self.assertEqual(manifest.get('prefix'), '/')
        self.assertEqual(manifest.get('endpoint'), self.command_url)
        self.assertTrue(any(item.get('name') == 'view' for item in manifest.get('commands') or []))

    def test_classroom_command_supports_view_alias_navigation(self):
        response = self._post_command('/shitu buju')
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        command_result = payload.get('command_result') or {}
        self.assertTrue(command_result.get('ok'))
        self.assertEqual(command_result.get('command'), 'view')
        self.assertEqual(command_result.get('subcommand'), 'layout')
        self.assertEqual((command_result.get('navigation') or {}).get('url'), reverse('layout_editor', args=[self.classroom.pk]))

    def test_classroom_command_can_swap_students(self):
        response = self._post_command('/zuowei jiaohuan 张三 李四')
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        command_result = payload.get('command_result') or {}
        self.assertTrue(command_result.get('ok'))
        self.assertEqual(command_result.get('kind'), 'mutation')
        self.assertTrue(command_result.get('needs_refresh'))

        self.student_a.refresh_from_db()
        self.student_b.refresh_from_db()
        self.assertEqual(self.student_a.assigned_seat.row, 1)
        self.assertEqual(self.student_a.assigned_seat.col, 2)
        self.assertEqual(self.student_b.assigned_seat.row, 1)
        self.assertEqual(self.student_b.assigned_seat.col, 1)

    def test_classroom_command_returns_student_info_with_default_query(self):
        response = self._post_command('/xuesheng 张三')
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        command_result = payload.get('command_result') or {}
        self.assertTrue(command_result.get('ok'))
        self.assertEqual(command_result.get('command'), 'students')
        self.assertEqual(command_result.get('subcommand'), 'info')
        self.assertEqual((command_result.get('data') or {}).get('name'), '张三')
        self.assertIn('学生：张三', payload.get('reply') or '')

    def test_classroom_command_lists_snapshots(self):
        LayoutSnapshot.objects.create(classroom=self.classroom, name='期中布局', data={'rows': 3, 'cols': 3})
        response = self._post_command('/kuaizhao liebiao')
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        command_result = payload.get('command_result') or {}
        self.assertTrue(command_result.get('ok'))
        items = (command_result.get('data') or {}).get('items') or []
        self.assertEqual(len(items), 1)
        self.assertEqual(items[0].get('name'), '期中布局')

    def test_ai_chat_uses_command_backend_for_slash_commands(self):
        with patch('seats.views._run_future_mode') as mock_run_future_mode:
            response = self.client.post(
                self.ai_chat_url,
                data=json.dumps({'action': 'message', 'message': '/view layout'}),
                content_type='application/json',
            )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload.get('status'), 'success')
        self.assertEqual((payload.get('command_result') or {}).get('command'), 'view')
        mock_run_future_mode.assert_not_called()

        conversation = self.classroom.ai_conversations.first()
        self.assertIsNotNone(conversation)
        messages = list(conversation.messages.order_by('created_at', 'pk'))
        self.assertEqual(len(messages), 2)
        self.assertEqual(messages[0].role, 'user')
        self.assertEqual(messages[0].content, '/view layout')
        self.assertEqual(messages[1].role, 'assistant')
        self.assertIn('布局视图', messages[1].content)

    def test_ai_chat_stream_returns_done_event_for_command(self):
        response = self.client.post(
            self.ai_chat_stream_url,
            data=json.dumps({'action': 'message', 'message': '/overview'}),
            content_type='application/json',
        )
        self.assertEqual(response.status_code, 200)
        streamed = ''.join(
            chunk if isinstance(chunk, str) else chunk.decode('utf-8')
            for chunk in response.streaming_content
        )
        self.assertIn('event: done', streamed)
        payload = json.loads(streamed.split('data: ', 1)[1].strip())
        self.assertEqual(payload.get('status'), 'success')
        self.assertEqual((payload.get('command_result') or {}).get('command'), 'overview')


class RuntimeReleaseManifestTests(unittest.TestCase):
    def test_get_current_version_prefers_runtime_release_manifest(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            runtime_manifest = root / 'runtime' / 'release.json'
            runtime_manifest.parent.mkdir(parents=True, exist_ok=True)
            runtime_manifest.write_text(
                json.dumps({'version': '2.3.4'}, ensure_ascii=False),
                encoding='utf-8',
            )

            with patch.object(desktop_runtime, 'iter_runtime_roots', return_value=[root]):
                with patch.dict('os.environ', {'FUCKSEATS_APP_VERSION': ''}, clear=False):
                    self.assertEqual(desktop_runtime.get_current_version(), '2.3.4')

    def test_get_current_version_prefers_env_over_manifest(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            runtime_manifest = root / 'runtime' / 'release.json'
            runtime_manifest.parent.mkdir(parents=True, exist_ok=True)
            runtime_manifest.write_text(
                json.dumps({'version': '2.3.4'}, ensure_ascii=False),
                encoding='utf-8',
            )

            with patch.object(desktop_runtime, 'iter_runtime_roots', return_value=[root]):
                with patch.dict('os.environ', {'FUCKSEATS_APP_VERSION': '9.9.9'}, clear=False):
                    self.assertEqual(desktop_runtime.get_current_version(), '9.9.9')


class PublicRedirectTests(TestCase):
    def test_api_json_redirects_to_app_manifest_api(self):
        response = self.client.get(reverse('app_manifest_redirect'))

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response['Location'], APP_MANIFEST_REDIRECT_URL)

    def test_api_json_redirect_preserves_query_string(self):
        response = self.client.get('/api.json?t=123&cache=no-store')

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response['Location'], f'{APP_MANIFEST_REDIRECT_URL}?t=123&cache=no-store')

    def test_update_txt_redirects_to_update_details_api(self):
        response = self.client.get(reverse('update_details_redirect'))

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response['Location'], UPDATE_DETAILS_REDIRECT_URL)

    def test_update_txt_redirect_preserves_query_string(self):
        response = self.client.get('/update.txt?t=123')

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response['Location'], f'{UPDATE_DETAILS_REDIRECT_URL}?t=123')


class SettingsPageVersionTests(TestCase):
    @patch('seats.context_processors.desktop_runtime.get_current_version', return_value='7.8.9')
    def test_settings_page_renders_current_version(self, _mock_get_current_version):
        response = self.client.get(reverse('settings'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, '7.8.9')

    @patch('seats.context_processors.desktop_runtime.get_current_version', return_value='7.8.9')
    def test_settings_page_contains_update_details_entry(self, _mock_get_current_version):
        response = self.client.get(reverse('settings'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'update-details-content')
        self.assertContains(response, '/update.txt')

class StudentTagSystemTests(TestCase):
    def setUp(self):
        self.classroom = Classroom.objects.create(name="标签测试班", rows=2, cols=2)
        self.alice = Student.objects.create(classroom=self.classroom, name="张三", student_id="T001", score=90)
        self.bob = Student.objects.create(classroom=self.classroom, name="李四", student_id="T002", score=80)

    def post_json(self, url, payload):
        return self.client.post(
            url,
            data=json.dumps(payload),
            content_type="application/json",
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

    def test_tag_crud_assignment_search_and_state_payload(self):
        create_response = self.post_json(
            reverse("student_tags", args=[self.classroom.pk]),
            {"name": "近视", "color": "0a59f7", "description": "需要靠前"},
        )
        self.assertEqual(create_response.status_code, 200)
        tag_id = create_response.json()["tag"]["id"]

        assign_response = self.post_json(
            reverse("assign_student_tags", args=[self.classroom.pk]),
            {"student_ids": [self.alice.pk], "tag_ids": [tag_id], "mode": "add"},
        )
        self.assertEqual(assign_response.status_code, 200)
        self.assertTrue(StudentTagMembership.objects.filter(student=self.alice, tag_id=tag_id).exists())

        search_response = self.client.get(reverse("search_students", args=[self.classroom.pk]), {"q": "近视"})
        self.assertEqual(search_response.status_code, 200)
        self.assertEqual([item["id"] for item in search_response.json()["students"]], [self.alice.pk])

        tag_search_response = self.client.get(
            reverse("search_students_by_tags", args=[self.classroom.pk]),
            {"tag_ids": str(tag_id)},
        )
        self.assertEqual(tag_search_response.status_code, 200)
        self.assertEqual([item["id"] for item in tag_search_response.json()["students"]], [self.alice.pk])

        state_response = self.client.get(reverse("classroom_state", args=[self.classroom.pk]))
        self.assertEqual(state_response.status_code, 200)
        state = state_response.json()
        self.assertEqual(state["tags"][0]["name"], "近视")
        self.assertIn("tag_rule_types", state)

    def test_tag_must_area_rule_affects_auto_arrange(self):
        tag = StudentTag.objects.create(classroom=self.classroom, name="近视")
        StudentTagMembership.objects.create(classroom=self.classroom, student=self.alice, tag=tag)
        StudentTagRule.objects.create(
            classroom=self.classroom,
            tag=tag,
            rule_type=StudentTagRule.RuleType.MUST_AREA,
            row_min=1,
            row_max=1,
            enabled=True,
        )

        response = self.client.post(
            reverse("auto_arrange_seats", args=[self.classroom.pk]),
            {"method": "random"},
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )
        self.assertEqual(response.status_code, 200, response.content.decode("utf-8"))
        self.alice.refresh_from_db()
        self.assertEqual(self.alice.assigned_seat.row, 1)

    def test_tag_rule_blocks_manual_move_outside_allowed_area(self):
        tag = StudentTag.objects.create(classroom=self.classroom, name="近视")
        StudentTagMembership.objects.create(classroom=self.classroom, student=self.alice, tag=tag)
        StudentTagRule.objects.create(
            classroom=self.classroom,
            tag=tag,
            rule_type=StudentTagRule.RuleType.MUST_AREA,
            row_min=1,
            row_max=1,
            enabled=True,
        )
        self.classroom.seats.filter(row=1, col=1).update(student=self.alice)
        self.classroom.seats.filter(row=1, col=2).update(student=self.bob)

        response = self.post_json(
            reverse("move_student", args=[self.classroom.pk]),
            {"student_id": self.alice.pk, "row": 2, "col": 1},
        )
        self.assertEqual(response.status_code, 400)
        self.assertIn("未坐在要求区域", response.json()["message"])
